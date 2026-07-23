import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import json
import pandas as pd
from datetime import datetime, timedelta
import os
from typing import Dict, List, Optional, Tuple

# Konstanten
DAYS_IN_PLANNING = 12
DAYS_PER_WEEK = 6
WORKING_DAYS = 10  # Nur Arbeitstage Mo-Fr (ohne Samstage)
WEEKDAY_NAMES = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag"]
WINDOW_GEOMETRY = "1400x900"
MIN_WINDOW_SIZE = (1200, 800)
CONFIG_FILE = "shift_config.json"
HISTORY_FILE = "shift_history.json"
ABSENCES_FILE = "absences.json"
CATCHUP_DELAY_DAYS = 2  # Tage bis Nachhol-Versuch für übersprungene Mitarbeiter
CATCHUP_PRIORITY_DAYS = 4  # Nach diesen Tagen bekommt Nachhol absolute Priorität
FAIR_MIN_GAP_SAME_SHIFT = 3  # Fair-Modus: min. Abstand (in Arbeitstagen) zwischen 2x derselben Schichtart

# Gewichtung der Soll-Verteilung: Einsätze von Pool-B-MA (brauchen Support) zählen
# in allen Fairness-Vergleichen 1,5-fach. Ergebnis bei üblicher Last (2-3 Einsätze
# pro 2-Wochen-Plan): Pool-C-MA tragen pro Plan ca. einen Einsatz mehr als Pool B.
# In der Hilfe wird das bewusst neutral als "Gewichtung nach Einsetzbarkeit" beschrieben.
FAIR_POOL_B_WEIGHT = 1.5

# Fairness-Horizont: wie weit zurück soll die Historie für die Fair-Verteilung berücksichtigt werden?
# Mapping Anzeige-Label -> Wochen rückwärts (0 = nur aktueller Plan, -1 = komplette Historie)
FAIRNESS_HORIZON_OPTIONS = [
    ("Nur dieser Plan", 0),
    ("1 Monat", 4),
    ("3 Monate", 13),
    ("6 Monate", 26),
    ("Komplett", -1),
]
FAIRNESS_HORIZON_DEFAULT_LABEL = "3 Monate"

# Excel Farbpalette
COLOR_GREEN = "A9D18E"
COLOR_YELLOW = "FFD966"
COLOR_GREY = "BFBFBF"

# Abwesenheits-Zeiträume: Label im GUI -> interner Schlüssel
# "vm" blockiert Vormittag UND Support (beides Vormittags-Einsätze), "nm" nur Nachmittag
ABSENCE_SCOPE_OPTIONS = [
    ("Ganztags", "ganz"),
    ("Nur vormittags", "vm"),
    ("Nur nachmittags", "nm"),
]
ABSENCE_SCOPE_LABELS = {key: label for label, key in ABSENCE_SCOPE_OPTIONS}

# Farben für die Plan-Prüfung (Treeview-Zeilen)
CHECK_COLOR_FEHLER = "#F4B6B6"   # rot: Regelverstoß
CHECK_COLOR_HINWEIS = "#FFE2B8"  # orange: Hinweis

# Schwellwert für Wochenbalance-Hinweis (max-min Einsätze innerhalb einer Woche)
WEEK_BALANCE_HINT_SPREAD = 3

# Max. Verbesserungs-Runden des Optimierungspasses im Fair-Modus
FAIR_IMPROVE_MAX_ROUNDS = 60


class ShiftPlanner:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Notdienst Schichtplaner v1.3")
        self.root.geometry(WINDOW_GEOMETRY)
        self.root.minsize(*MIN_WINDOW_SIZE)

        # Konfigurationsdatei
        self.config_file = CONFIG_FILE

        # Standardkonfiguration
        self.config: Dict[str, List[str]] = {
            "pool_vm_alle": [],            # Vormittag - können alles
            "pool_vm_teilweise": [],       # Vormittag - können nicht alles (brauchen Support)
            "pool_vm_support": [],         # Vormittag - Support für Pool B
            "pool_nm_alle": [],            # Nachmittag - können alles
            "pool_freitag_abwesend": [],   # Freitags nicht verfügbar
            "pool_mo_mi_abwesend": [],     # Montag/Mittwoch nicht verfügbar
            "feiertage": []                # [{datum: "25.12", name: "1. Weihnachtstag", mitarbeiter: "XX"}]
        }

        # Cache für Performance
        self._all_employees_cache: Optional[List[str]] = None
        self._cache_dirty = True

        # Interne Zustandsverwaltung
        self.planning_result: List[Dict[str, str]] = []
        # Abwesenheiten: Liste von {"ma": str, "von": datetime, "bis": datetime, "zeit": "ganz"|"vm"|"nm"}
        self.absences: List[Dict] = []
        self.saved_plans: List[Dict] = []  # Gespeicherte Pläne für Auswertung
        self.manual_stats_corrections: Dict[str, Dict[str, int]] = {} # Manuelle Korrekturen {Name: {VM: +1, ...}}
        self.history_file = HISTORY_FILE
        self._plan_manually_edited = False  # Flag: Plan wurde nachträglich geändert
        self._plan_violations: List[Dict] = []  # Ergebnisse der letzten Plan-Prüfung
        self._fairness_base_counts: Optional[Dict[str, Dict[str, int]]] = None  # Historie-Stand vor aktuellem Plan

        self.load_config()
        self.load_history()
        self._load_absences()
        self.create_gui()

        # Automatisches Laden der Mitarbeiterliste nach GUI-Erstellung
        self.root.after(100, self._auto_update_employee_list)
        # Startdatum automatisch vorbefüllen (nächster Montag nach letztem Plan)
        self.root.after(150, self._prefill_start_date)

    # -------------------- Konfiguration laden/speichern --------------------

    def load_config(self) -> None:
        """Lädt Konfiguration aus JSON-Datei"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    loaded_config = json.load(f)
                    # Sicherstellen, dass alle benötigten Pools existieren
                    for key in self.config.keys():
                        if key not in loaded_config:
                            loaded_config[key] = []
                    self.config = loaded_config
                    self._cache_dirty = True
        except json.JSONDecodeError as e:
            messagebox.showerror("Fehler", f"Ungültige JSON-Datei: {e}")
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Laden der Konfiguration: {e}")

    def save_config(self) -> None:
        """Speichert Konfiguration in JSON-Datei"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
            self._cache_dirty = True
            messagebox.showinfo("Erfolg", "Konfiguration gespeichert!")
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Speichern: {e}")

    def load_history(self) -> None:
        """Lädt gespeicherte Pläne aus History-Datei"""
        try:
            if os.path.exists(self.history_file):
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.saved_plans = data.get("plans", [])
                    self.manual_stats_corrections = data.get("manual_corrections", {})
                    self._eval_filter_from = data.get("filter_from", "")
                    self._eval_filter_to = data.get("filter_to", "")
                    self._fairness_horizon_label = data.get("fairness_horizon_label", FAIRNESS_HORIZON_DEFAULT_LABEL)
            else:
                self.saved_plans = []
                self.manual_stats_corrections = {}
                self._eval_filter_from = ""
                self._eval_filter_to = ""
                self._fairness_horizon_label = FAIRNESS_HORIZON_DEFAULT_LABEL
        except Exception:
            self.saved_plans = []
            self.manual_stats_corrections = {}
            self._eval_filter_from = ""
            self._eval_filter_to = ""
            self._fairness_horizon_label = FAIRNESS_HORIZON_DEFAULT_LABEL

    def save_history(self) -> None:
        """Speichert Pläne in History-Datei"""
        try:
            data = {
                "plans": self.saved_plans,
                "manual_corrections": getattr(self, 'manual_stats_corrections', {}),
                "filter_from": getattr(self, '_eval_filter_from', ""),
                "filter_to": getattr(self, '_eval_filter_to', ""),
                "fairness_horizon_label": getattr(self, '_fairness_horizon_label', FAIRNESS_HORIZON_DEFAULT_LABEL)
            }
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Speichern der Historie: {e}")


    def _get_last_tag10_data(self) -> Optional[Dict[str, str]]:
        """
        Holt Daten basierend auf letztem Plan:
        - Datum: Letzter Eintrag (Fr) + 3 Tage = Nächster Montag
        - MA: Rotation wird FORTGESETZT basierend auf letztem Do (vorletzter Arbeitstag)
        - Support: Tag 1 (Mo) erzwingt Support (wie Montag Woche 2)
        Unterstützt sowohl alte 12-Tage-Pläne (mit Sa) als auch neue 10-Tage-Pläne (ohne Sa).
        """
        if not self.saved_plans:
            return None
        
        try:
            # 1. Letzten Plan finden
            sorted_plans = sorted(
                self.saved_plans,
                key=lambda p: datetime.strptime(p.get("start_date", "01.01.1900"), "%d.%m.%Y"),
                reverse=True
            )
            last_plan = sorted_plans[0]
            entries = last_plan.get("entries", [])
            
            if len(entries) < 10:
                return None

            # 2. Datum berechnen: Letzter Eintrag → nächsten Montag finden
            last_entry = entries[-1]
            last_date = datetime.strptime(last_entry.get("Datum", ""), "%d.%m.%Y")
            # Nächsten Montag berechnen (egal ob letzter Tag Fr oder Sa war)
            days_until_monday = (7 - last_date.weekday()) % 7
            if days_until_monday == 0:
                days_until_monday = 7  # Falls letzter Tag selbst ein Montag wäre
            new_start_date = last_date + timedelta(days=days_until_monday)
            
            # 3. Letzten Donnerstag im Plan finden (vorletzter Arbeitstag)
            #    Bei 10 Einträgen: Index 8 = Do Woche 2
            #    Bei 12 Einträgen (alt): Index 9 = Do Woche 2 (Sa bei 5 und 11)
            ref_entry = None
            for i in range(len(entries) - 1, -1, -1):
                wochentag = entries[i].get("Wochentag", "")
                if wochentag.startswith("Do"):
                    ref_entry = entries[i]
                    break
            
            if ref_entry is None:
                # Fallback: vorletzter Eintrag
                ref_entry = entries[-2] if len(entries) >= 2 else entries[-1]
            
            prev_vm = ref_entry.get("Vormittag", "")
            prev_nm = ref_entry.get("Nachmittag", "")
            prev_support = ref_entry.get("Support", "")

            # Falls Support leer war, rückwärts suchen für Rotation
            if not prev_support:
                for i in range(len(entries) - 1, -1, -1):
                    s = entries[i].get("Support", "")
                    if s:
                        prev_support = s
                        break
            
            # 4. Nachfolger bestimmen
            def get_successor(name: str, pool_key: str) -> str:
                pool = self.config.get(pool_key, [])
                if not pool: return ""
                if name in pool:
                    idx = pool.index(name)
                    return pool[(idx + 1) % len(pool)]
                return pool[0] # Fallback: Erster der Liste

            next_vm = get_successor(prev_vm, "pool_vm_alle")
            next_nm = get_successor(prev_nm, "pool_nm_alle")
            # Support ist für Tag 1 (Montag) PFLICHT ("wie Montag Woche 2")
            next_support = get_successor(prev_support, "pool_vm_support")

            return {
                "start_date": new_start_date.strftime("%d.%m.%Y"),
                "vm": next_vm,
                "nm": next_nm,
                "support": next_support
            }
            
        except Exception:
            pass
        
        return None



    def _on_auto_mode_toggle(self) -> None:
        """Handler für Automatik-Modus Toggle"""
        if self.auto_mode_var.get():
            # Automatik aktiviert: Felder aus letztem Tag 10 befüllen
            tag10_data = self._get_last_tag10_data()
            
            if tag10_data:
                # Felder leeren und neu befüllen
                self.start_date_entry.delete(0, tk.END)
                self.start_date_entry.insert(0, tag10_data["start_date"])
                
                self.first_vm_entry.delete(0, tk.END)
                self.first_vm_entry.insert(0, tag10_data["vm"])
                
                self.first_nm_entry.delete(0, tk.END)
                self.first_nm_entry.insert(0, tag10_data["nm"])
                
                self.first_support_entry.delete(0, tk.END)
                self.first_support_entry.insert(0, tag10_data["support"])
            else:
                messagebox.showinfo(
                    "Automatik-Modus",
                    "Keine gespeicherten Pläne gefunden.\n"
                    "Bitte zuerst einen Plan speichern oder manuell eingeben."
                )
                self.auto_mode_var.set(False)

    def _on_fair_mode_toggle(self) -> None:
        """Handler für Fair-Mode-Toggle: Aktiviert/deaktiviert das Horizont-Dropdown."""
        if hasattr(self, 'fairness_horizon_combo'):
            new_state = "readonly" if self.fair_mode_var.get() else "disabled"
            try:
                self.fairness_horizon_combo.configure(state=new_state)
            except tk.TclError:
                pass

    def _on_fairness_horizon_change(self) -> None:
        """Persistiert die Auswahl des Fairness-Horizonts."""
        try:
            self._fairness_horizon_label = self.fairness_horizon_var.get()
            self.save_history()
        except Exception:
            pass

    def _update_fairness_metric_display(self) -> None:
        """Aktualisiert das Fairness-Label unter dem Plan basierend auf dem zuletzt erzeugten Plan."""
        if not hasattr(self, 'fairness_metric_label'):
            return
        counts = getattr(self, '_fairness_after', None)
        if not counts:
            self.fairness_metric_label.config(
                text="Fairness: – (Plan erzeugen, um Metrik zu sehen)"
            )
            return
        metrics = self._compute_fairness_metrics(counts)
        if not metrics:
            self.fairness_metric_label.config(text="Fairness: keine Daten")
            return
        sigma = metrics["sigma"]
        baseline = getattr(self, '_fairness_baseline', None)
        baseline_text = ""
        if baseline:
            base_metrics = self._compute_fairness_metrics(baseline)
            if base_metrics:
                delta = sigma - base_metrics["sigma"]
                arrow = "↓" if delta < -0.01 else ("↑" if delta > 0.01 else "→")
                baseline_text = f"  (vorher σ={base_metrics['sigma']:.2f} {arrow})"
        text = (
            f"Fairness: σ={sigma:.2f}  "
            f"Min: {metrics['min_ma']} ({metrics['min_value']})  "
            f"Max: {metrics['max_ma']} ({metrics['max_value']})  "
            f"Spread: {metrics['spread']}{baseline_text}"
        )
        self.fairness_metric_label.config(text=text)

    def _get_all_employees(self) -> List[str]:
        """Gibt gecachte Liste aller Mitarbeiter zurück (Performance-Optimierung)"""
        if self._cache_dirty or self._all_employees_cache is None:
            all_employees = set()
            for pool_key in ["pool_vm_alle", "pool_vm_teilweise", "pool_vm_support", "pool_nm_alle"]:
                all_employees.update(self.config[pool_key])
            self._all_employees_cache = sorted(list(all_employees))
            self._cache_dirty = False
        return self._all_employees_cache

    def _parse_employee_input(self, input_str: str) -> List[str]:
        """Parst Mitarbeiter-Eingabe (kommagetrennt) mit Validierung"""
        return [x.strip() for x in input_str.split(",") if x.strip()]

    def _auto_update_employee_list(self) -> None:
        """Automatisches Laden der Mitarbeiterliste (ohne Messagebox)"""
        try:
            all_employees = self._get_all_employees()
            if hasattr(self, 'employee_combo'):
                self.employee_combo['values'] = all_employees
            if hasattr(self, 'holiday_employee_combo'):
                self.holiday_employee_combo['values'] = ['-- Kein Notdienst --'] + all_employees
        except Exception:
            pass  # Stilles Fehlschlagen beim automatischen Update

    # -------------------- GUI --------------------

    def create_gui(self):
        """Erstellt die Benutzeroberfläche"""
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        # Tab 1: Pool-Konfiguration
        self.create_pool_config_tab(notebook)
        # Tab 2: Schichtplanung
        self.create_shift_planning_tab(notebook)
        # Tab 3: Auswertung
        self.create_evaluation_tab(notebook)
        # Tab 4: Historie
        self.create_history_tab(notebook)
        # Tab 5: Hilfe
        self.create_help_tab(notebook)

    def _create_pool_entry(self, parent: ttk.Frame, row: int, label_text: str,
                          config_key: str, **label_kwargs) -> tk.Entry:
        """Hilfsmethode zum Erstellen eines Pool-Entry-Feldes (DRY)"""
        ttk.Label(parent, text=label_text, **label_kwargs).grid(
            row=row, column=0, sticky="w", padx=5, pady=5
        )
        entry = tk.Entry(parent, width=60)
        entry.grid(row=row, column=1, padx=5, pady=5)
        entry.insert(0, ",".join(self.config[config_key]))
        return entry

    def create_pool_config_tab(self, notebook: ttk.Notebook) -> None:
        """Erstellt Tab für Pool-Konfiguration mit Feiertagen"""
        pool_frame = ttk.Frame(notebook, padding=10)
        notebook.add(pool_frame, text="Pool-Konfiguration")

        # Hauptcontainer: Links Pools, Rechts Feiertage
        main_container = ttk.PanedWindow(pool_frame, orient="horizontal")
        main_container.pack(fill="both", expand=True)

        # === LINKER BEREICH: Pool-Konfiguration ===
        left_frame = ttk.LabelFrame(main_container, text=" Mitarbeiter-Pools ", padding=10)

        # Pools erstellen (DRY mit Hilfsmethode)
        self.pool_vm_alle_entry = self._create_pool_entry(
            left_frame, 0, "Pool A - Vormittag (können alles):", "pool_vm_alle"
        )
        self.pool_vm_teilweise_entry = self._create_pool_entry(
            left_frame, 1, "Pool B - Vormittag (brauchen Unterstützung):", "pool_vm_teilweise"
        )
        self.pool_vm_support_entry = self._create_pool_entry(
            left_frame, 2, "Pool C - Support für Pool B:", "pool_vm_support"
        )
        self.pool_nm_alle_entry = self._create_pool_entry(
            left_frame, 3, "Pool D - Nachmittag (können alles):", "pool_nm_alle"
        )
        self.pool_freitag_abwesend_entry = self._create_pool_entry(
            left_frame, 4, "Pool E - Freitags NICHT verfügbar:", "pool_freitag_abwesend",
            font=('TkDefaultFont', 9, 'bold'), foreground='orange'
        )
        self.pool_mo_mi_abwesend_entry = self._create_pool_entry(
            left_frame, 5, "Pool F - Mo/Mi NICHT verfügbar:", "pool_mo_mi_abwesend",
            font=('TkDefaultFont', 9, 'bold'), foreground='#e67300'
        )

        # Speichern Button
        ttk.Button(left_frame, text="Pools speichern", command=self.save_pools).grid(
            row=6, column=1, pady=15, sticky="e"
        )

        info_text = """Hinweise:
• Mitarbeiter mit Komma trennen (z.B. RR,AN,MH)
• Reihenfolge wird bei der Planung berücksichtigt"""
        ttk.Label(left_frame, text=info_text, justify="left", font=('TkDefaultFont', 8),
                 foreground='gray').grid(row=7, column=0, columnspan=2, padx=5, pady=5, sticky="w")

        main_container.add(left_frame, weight=2)

        # === RECHTER BEREICH: Feiertage ===
        right_frame = ttk.LabelFrame(main_container, text=" Feiertage (jährlich) ", padding=10)

        # Eingabefelder für neuen Feiertag
        input_frame = ttk.Frame(right_frame)
        input_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(input_frame, text="Datum (TT.MM):").grid(row=0, column=0, sticky="w", padx=2)
        self.holiday_date_var = tk.StringVar()
        self.holiday_date_entry = ttk.Entry(input_frame, textvariable=self.holiday_date_var, width=8)
        self.holiday_date_entry.grid(row=0, column=1, padx=5)

        ttk.Label(input_frame, text="Name:").grid(row=0, column=2, sticky="w", padx=2)
        self.holiday_name_var = tk.StringVar()
        self.holiday_name_entry = ttk.Entry(input_frame, textvariable=self.holiday_name_var, width=20)
        self.holiday_name_entry.grid(row=0, column=3, padx=5)

        input_frame2 = ttk.Frame(right_frame)
        input_frame2.pack(fill="x", pady=(0, 10))

        ttk.Label(input_frame2, text="Notdienst MA:").grid(row=0, column=0, sticky="w", padx=2)
        self.holiday_employee_var = tk.StringVar()
        self.holiday_employee_combo = ttk.Combobox(input_frame2, textvariable=self.holiday_employee_var, width=8)
        self.holiday_employee_combo.grid(row=0, column=1, padx=5)

        # Buttons
        btn_frame = ttk.Frame(right_frame)
        btn_frame.pack(fill="x", pady=(0, 10))
        ttk.Button(btn_frame, text="+ Hinzufügen", command=self.add_holiday, width=12).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="- Entfernen", command=self.remove_holiday, width=12).pack(side="left", padx=2)

        # Feiertage-Liste
        holiday_columns = ("Datum", "Name", "MA")
        self.holiday_tree = ttk.Treeview(right_frame, columns=holiday_columns, show="headings", height=12)
        self.holiday_tree.heading("Datum", text="Datum")
        self.holiday_tree.heading("Name", text="Name")
        self.holiday_tree.heading("MA", text="Notdienst")
        self.holiday_tree.column("Datum", width=70, anchor="center")
        self.holiday_tree.column("Name", width=150, anchor="w")
        self.holiday_tree.column("MA", width=70, anchor="center")

        scrollbar_holiday = ttk.Scrollbar(right_frame, orient="vertical", command=self.holiday_tree.yview)
        self.holiday_tree.configure(yscrollcommand=scrollbar_holiday.set)

        self.holiday_tree.pack(side="left", fill="both", expand=True)
        scrollbar_holiday.pack(side="right", fill="y")

        # Feiertage aus Config laden
        self.update_holiday_display()

        main_container.add(right_frame, weight=1)

        info_label = ttk.Label(pool_frame, text="Feiertage werden dauerhaft gespeichert. "
                              "Der Notdienst-MA wird in der Planungswoche von regulären Schichten ausgeschlossen.",
                              font=('TkDefaultFont', 8), foreground='gray')
        info_label.pack(fill="x", pady=(10, 0))

    def create_shift_planning_tab(self, notebook: ttk.Notebook) -> None:
        """Erstellt Tab für Schichtplanung - MODERNES LAYOUT"""
        planning_frame = ttk.Frame(notebook, padding=10)
        notebook.add(planning_frame, text="Schichtplanung")

        # Modernes Styling konfigurieren
        self._configure_modern_styles()

        # === OBERER BEREICH: Grundeinstellungen (kompakt) ===
        settings_frame = ttk.LabelFrame(planning_frame, text=" Grundeinstellungen ", padding=10)
        settings_frame.pack(fill="x", pady=(0, 10))

        # Eingabefelder in einer Zeile
        input_frame = ttk.Frame(settings_frame)
        input_frame.pack(fill="x")

        # Startdatum
        ttk.Label(input_frame, text="Startdatum:").grid(row=0, column=0, sticky="w", padx=(0, 5))
        self.start_date_entry = ttk.Entry(input_frame, width=14)
        self.start_date_entry.grid(row=0, column=1, padx=(0, 15))

        # 1. Tag VM
        ttk.Label(input_frame, text="1. Tag VM:").grid(row=0, column=2, sticky="w", padx=(0, 5))
        self.first_vm_entry = ttk.Entry(input_frame, width=8)
        self.first_vm_entry.grid(row=0, column=3, padx=(0, 15))

        # 1. Tag NM
        ttk.Label(input_frame, text="1. Tag NM:").grid(row=0, column=4, sticky="w", padx=(0, 5))
        self.first_nm_entry = ttk.Entry(input_frame, width=8)
        self.first_nm_entry.grid(row=0, column=5, padx=(0, 15))

        # 1. Tag Support
        ttk.Label(input_frame, text="Support:").grid(row=0, column=6, sticky="w", padx=(0, 5))
        self.first_support_entry = ttk.Entry(input_frame, width=8)
        self.first_support_entry.grid(row=0, column=7, padx=(0, 20))

        # Buttons
        ttk.Button(input_frame, text="Planung erstellen", command=self.create_planning,
                  style="Accent.TButton").grid(row=0, column=8, padx=5)
        ttk.Button(input_frame, text="Plan speichern", command=self.save_current_plan).grid(row=0, column=9, padx=5)
        ttk.Button(input_frame, text="Excel Export", command=self.export_excel).grid(row=0, column=10, padx=5)

        # Optionen (zweite Zeile)
        auto_frame = ttk.Frame(settings_frame)
        auto_frame.pack(fill="x", pady=(10, 0))
        
        self.auto_mode_var = tk.BooleanVar(value=False)
        self.auto_mode_check = ttk.Checkbutton(
            auto_frame, 
            text="Automatik-Modus (übernimmt Werte vom letzten gespeicherten Tag 10)",
            variable=self.auto_mode_var,
            command=self._on_auto_mode_toggle
        )
        self.auto_mode_check.pack(side="left")

        self.monday_normal_var = tk.BooleanVar(value=False)
        self.monday_normal_check = ttk.Checkbutton(
            auto_frame, 
            text="Montag = normaler Tag (kein Extra-Support)",
            variable=self.monday_normal_var
        )
        self.monday_normal_check.pack(side="left", padx=(20, 0))

        self.fair_mode_var = tk.BooleanVar(value=True)
        self.fair_mode_check = ttk.Checkbutton(
            auto_frame,
            text="Faire Verteilung (zählbasiert)",
            variable=self.fair_mode_var,
            command=self._on_fair_mode_toggle
        )
        self.fair_mode_check.pack(side="left", padx=(20, 0))

        # Fairness-Horizont (nur aktiv wenn Fair-Mode an)
        ttk.Label(auto_frame, text="Historie:").pack(side="left", padx=(20, 5))
        self.fairness_horizon_var = tk.StringVar(
            value=getattr(self, '_fairness_horizon_label', FAIRNESS_HORIZON_DEFAULT_LABEL)
        )
        self.fairness_horizon_combo = ttk.Combobox(
            auto_frame,
            textvariable=self.fairness_horizon_var,
            values=[label for label, _ in FAIRNESS_HORIZON_OPTIONS],
            state="readonly",
            width=14,
        )
        self.fairness_horizon_combo.pack(side="left")
        self.fairness_horizon_combo.bind("<<ComboboxSelected>>", lambda e: self._on_fairness_horizon_change())

        # === HAUPTBEREICH: Planungsergebnis (links) + Abwesenheiten (rechts) ===
        main_paned = ttk.PanedWindow(planning_frame, orient="horizontal")
        main_paned.pack(fill="both", expand=True)

        # === LINKER BEREICH: Planungsergebnis (größer) ===
        result_frame = ttk.LabelFrame(main_paned, text=" Planungsergebnis ", padding=10)

        tree_container = ttk.Frame(result_frame)
        tree_container.pack(side="top", fill="both", expand=True)

        columns = ("Datum", "Tag", "VM", "NM", "Support")
        self.result_tree = ttk.Treeview(tree_container, columns=columns, show="headings",
                                        height=14, style="Modern.Treeview")

        # Spalten-Konfiguration (kompakter)
        column_config = {
            "Datum": (85, "center"),
            "Tag": (80, "center"),
            "VM": (60, "center"),
            "NM": (60, "center"),
            "Support": (60, "center")
        }
        for col_name, (width, anchor) in column_config.items():
            self.result_tree.heading(col_name, text=col_name)
            self.result_tree.column(col_name, width=width, minwidth=50, anchor=anchor)

        scrollbar_result = ttk.Scrollbar(tree_container, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=scrollbar_result.set)

        self.result_tree.pack(side="left", fill="both", expand=True)
        scrollbar_result.pack(side="right", fill="y")

        # Zeilen-Färbung für die automatische Plan-Prüfung
        self.result_tree.tag_configure("check_fehler", background=CHECK_COLOR_FEHLER)
        self.result_tree.tag_configure("check_hinweis", background=CHECK_COLOR_HINWEIS)

        # Double-click zum Bearbeiten
        self.result_tree.bind("<Double-1>", self._on_result_double_click)
        self._edit_entry = None  # Aktives Eingabefeld

        # === Plan-Prüfung: automatische Regelkontrolle unter dem Plan ===
        check_frame = ttk.LabelFrame(result_frame, text=" Plan-Prüfung ", padding=5)
        check_frame.pack(side="bottom", fill="x", pady=(8, 0))

        self.check_status_label = ttk.Label(check_frame, text="Noch kein Plan erstellt.",
                                            font=('Segoe UI', 9, 'bold'))
        self.check_status_label.pack(anchor="w")

        check_columns = ("Schwere", "Meldung")
        self.check_tree = ttk.Treeview(check_frame, columns=check_columns, show="headings",
                                       height=4, style="Modern.Treeview")
        self.check_tree.heading("Schwere", text="Schwere")
        self.check_tree.heading("Meldung", text="Meldung")
        self.check_tree.column("Schwere", width=70, anchor="center")
        self.check_tree.column("Meldung", width=500, anchor="w")
        self.check_tree.tag_configure("fehler", background=CHECK_COLOR_FEHLER)
        self.check_tree.tag_configure("hinweis", background=CHECK_COLOR_HINWEIS)

        scrollbar_check = ttk.Scrollbar(check_frame, orient="vertical", command=self.check_tree.yview)
        self.check_tree.configure(yscrollcommand=scrollbar_check.set)
        self.check_tree.pack(side="left", fill="both", expand=True)
        scrollbar_check.pack(side="right", fill="y")

        main_paned.add(result_frame, weight=3)

        # Fairness-Metrik-Anzeige unter dem Plan (am Boden gepinnt, damit es nicht von main_paned verdrängt wird)
        fairness_frame = ttk.Frame(planning_frame)
        fairness_frame.pack(side="bottom", fill="x", pady=(8, 0))
        self.fairness_metric_label = ttk.Label(
            fairness_frame,
            text="Fairness: – (Plan erzeugen, um Metrik zu sehen)",
            font=('Segoe UI', 9, 'italic'),
        )
        self.fairness_metric_label.pack(side="left")

        # === RECHTER BEREICH: Abwesenheiten (oben) + Wochenlast (unten) ===
        right_container = ttk.Frame(main_paned)

        absence_frame = ttk.LabelFrame(right_container, text=" Abwesenheiten / Termine ", padding=10)
        absence_frame.pack(side="top", fill="both", expand=True)

        input_absence_frame = ttk.Frame(absence_frame)
        input_absence_frame.pack(fill="x", pady=(0, 5))

        ttk.Label(input_absence_frame, text="MA:").grid(row=0, column=0, sticky="w")
        self.employee_var = tk.StringVar()
        self.employee_combo = ttk.Combobox(input_absence_frame, textvariable=self.employee_var, width=10)
        self.employee_combo.grid(row=0, column=1, padx=5, sticky="w")

        ttk.Label(input_absence_frame, text="Zeit:").grid(row=0, column=2, sticky="w")
        self.absence_scope_var = tk.StringVar(value=ABSENCE_SCOPE_OPTIONS[0][0])
        self.absence_scope_combo = ttk.Combobox(
            input_absence_frame, textvariable=self.absence_scope_var, width=14,
            values=[label for label, _ in ABSENCE_SCOPE_OPTIONS], state="readonly"
        )
        self.absence_scope_combo.grid(row=0, column=3, padx=5, sticky="w")

        ttk.Label(input_absence_frame, text="Von:").grid(row=1, column=0, sticky="w", pady=(5, 0))
        self.absence_from_var = tk.StringVar()
        self.absence_from_combo = ttk.Combobox(input_absence_frame, textvariable=self.absence_from_var, width=12)
        self.absence_from_combo.grid(row=1, column=1, padx=5, pady=(5, 0), sticky="w")

        ttk.Label(input_absence_frame, text="Bis:").grid(row=1, column=2, sticky="w", pady=(5, 0))
        self.absence_to_var = tk.StringVar()
        self.absence_to_combo = ttk.Combobox(input_absence_frame, textvariable=self.absence_to_var, width=12)
        self.absence_to_combo.grid(row=1, column=3, padx=5, pady=(5, 0), sticky="w")

        hint = ttk.Label(absence_frame, text="Datum: TT.MM.JJJJ · 'Bis' leer = ein Tag · mehrere MA mit Komma · bleibt gespeichert",
                         font=('TkDefaultFont', 8), foreground='gray')
        hint.pack(fill="x", pady=(0, 5))

        btn_frame = ttk.Frame(absence_frame)
        btn_frame.pack(fill="x", pady=(0, 10))
        ttk.Button(btn_frame, text="+ Hinzufügen", command=self.add_absence, width=12).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="- Entfernen", command=self.remove_absence, width=12).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="Laden", command=self.update_employee_list, width=8).pack(side="left", padx=2)

        absence_columns = ("MA", "Von", "Bis", "Zeit")
        self.absence_tree = ttk.Treeview(absence_frame, columns=absence_columns, show="headings",
                                         height=8, style="Modern.Treeview")
        self.absence_tree.heading("MA", text="MA")
        self.absence_tree.heading("Von", text="Von")
        self.absence_tree.heading("Bis", text="Bis")
        self.absence_tree.heading("Zeit", text="Zeit")
        self.absence_tree.column("MA", width=50, anchor="center")
        self.absence_tree.column("Von", width=80, anchor="center")
        self.absence_tree.column("Bis", width=80, anchor="center")
        self.absence_tree.column("Zeit", width=100, anchor="center")

        scrollbar_absence = ttk.Scrollbar(absence_frame, orient="vertical", command=self.absence_tree.yview)
        self.absence_tree.configure(yscrollcommand=scrollbar_absence.set)
        self.absence_tree.pack(side="left", fill="both", expand=True)
        scrollbar_absence.pack(side="right", fill="y")

        # === Wochenlast: Einsätze pro MA im aktuellen Plan ===
        week_frame = ttk.LabelFrame(right_container, text=" Wochenlast (aktueller Plan) ", padding=10)
        week_frame.pack(side="bottom", fill="both", pady=(8, 0))

        week_columns = ("MA", "W1", "W2", "Ges")
        self.week_load_tree = ttk.Treeview(week_frame, columns=week_columns, show="headings",
                                           height=9, style="Modern.Treeview")
        for col, width in [("MA", 60), ("W1", 50), ("W2", 50), ("Ges", 50)]:
            self.week_load_tree.heading(col, text=col)
            self.week_load_tree.column(col, width=width, anchor="center")
        self.week_load_tree.tag_configure("max_load", background=CHECK_COLOR_HINWEIS)
        self.week_load_tree.tag_configure("min_load", background="#D6EFC7")

        scrollbar_week = ttk.Scrollbar(week_frame, orient="vertical", command=self.week_load_tree.yview)
        self.week_load_tree.configure(yscrollcommand=scrollbar_week.set)
        self.week_load_tree.pack(side="left", fill="both", expand=True)
        scrollbar_week.pack(side="right", fill="y")

        main_paned.add(right_container, weight=1)

        # Gespeicherte Abwesenheiten direkt anzeigen
        self.update_absence_display()

        # Initialen State des Horizont-Dropdowns an Fair-Mode-Toggle koppeln
        self._on_fair_mode_toggle()

    def _configure_modern_styles(self) -> None:
        """Konfiguriert moderne ttk Styles"""
        style = ttk.Style()

        # Versuche ein moderneres Theme zu verwenden
        available_themes = style.theme_names()
        if 'clam' in available_themes:
            style.theme_use('clam')
        elif 'vista' in available_themes:
            style.theme_use('vista')

        # Moderne Treeview-Styles
        style.configure("Modern.Treeview",
                       rowheight=25,
                       font=('Segoe UI', 9))
        style.configure("Modern.Treeview.Heading",
                       font=('Segoe UI', 9, 'bold'),
                       padding=5)

        # Accent Button Style
        style.configure("Accent.TButton",
                       font=('Segoe UI', 9, 'bold'))

        # LabelFrame Style
        style.configure("TLabelframe.Label",
                       font=('Segoe UI', 10, 'bold'))

    def create_evaluation_tab(self, notebook: ttk.Notebook) -> None:
        """Erstellt Tab für Auswertung der Notdienste"""
        eval_frame = ttk.Frame(notebook, padding=10)
        notebook.add(eval_frame, text="Auswertung")

        # === FILTERBEREICH ===
        filter_frame = ttk.LabelFrame(eval_frame, text=" Datumsfilter ", padding=10)
        filter_frame.pack(fill="x", pady=(0, 10))

        filter_input = ttk.Frame(filter_frame)
        filter_input.pack(fill="x")

        ttk.Label(filter_input, text="Von (TT.MM.YYYY):").grid(row=0, column=0, sticky="w", padx=(0, 5))
        self.eval_from_var = tk.StringVar(value=getattr(self, '_eval_filter_from', ''))
        self.eval_from_entry = ttk.Entry(filter_input, textvariable=self.eval_from_var, width=14)
        self.eval_from_entry.grid(row=0, column=1, padx=(0, 15))

        ttk.Label(filter_input, text="Bis (TT.MM.YYYY):").grid(row=0, column=2, sticky="w", padx=(0, 5))
        self.eval_to_var = tk.StringVar(value=getattr(self, '_eval_filter_to', ''))
        self.eval_to_entry = ttk.Entry(filter_input, textvariable=self.eval_to_var, width=14)
        self.eval_to_entry.grid(row=0, column=3, padx=(0, 15))

        ttk.Button(filter_input, text="Aktualisieren", command=self.update_evaluation_display).grid(row=0, column=4, padx=5)
        ttk.Button(filter_input, text="Filter speichern", command=self._save_filter).grid(row=0, column=5, padx=5)

        # === STATISTIK-ANZEIGE ===
        stats_frame = ttk.LabelFrame(eval_frame, text=" Statistik pro Mitarbeiter ", padding=10)
        stats_frame.pack(fill="both", expand=True)

        columns = ("MA", "VM", "NM", "Support", "Gesamt")
        self.stats_tree = ttk.Treeview(stats_frame, columns=columns, show="headings",
                                       height=20, style="Modern.Treeview")

        column_config = {
            "MA": (100, "center"),
            "VM": (80, "center"),
            "NM": (80, "center"),
            "Support": (80, "center"),
            "Gesamt": (80, "center")
        }
        for col_name, (width, anchor) in column_config.items():
            self.stats_tree.heading(col_name, text=col_name)
            self.stats_tree.column(col_name, width=width, minwidth=60, anchor=anchor)

        scrollbar_stats = ttk.Scrollbar(stats_frame, orient="vertical", command=self.stats_tree.yview)
        self.stats_tree.configure(yscrollcommand=scrollbar_stats.set)

        self.stats_tree.pack(side="left", fill="both", expand=True)
        scrollbar_stats.pack(side="right", fill="y")

        # Double-click zum Bearbeiten der Statistik
        self.stats_tree.bind("<Double-1>", self._on_stats_double_click)
        self._stats_edit_entry = None

        # === INFO-BEREICH ===
        info_frame = ttk.Frame(eval_frame)
        info_frame.pack(fill="x", pady=(10, 0))

        self.stats_info_label = ttk.Label(info_frame, text="Keine Pläne gespeichert.", font=('Segoe UI', 9))
        self.stats_info_label.pack(side="left")

        ttk.Button(info_frame, text="Alle Pläne löschen", command=self._clear_all_plans).pack(side="right")

        # Initial aktualisieren
        self.root.after(200, self.update_evaluation_display)

    # -------------------- Historie --------------------

    def create_history_tab(self, notebook: ttk.Notebook) -> None:
        """Erstellt Tab für historische Planungsübersicht"""
        history_frame = ttk.Frame(notebook, padding=10)
        notebook.add(history_frame, text="Historie")

        # === OBERER BEREICH: Planübersicht ===
        overview_frame = ttk.LabelFrame(history_frame, text=" Gespeicherte Pläne ", padding=10)
        overview_frame.pack(fill="x", pady=(0, 10))

        # Treeview für Planübersicht
        hist_columns = ("Zeitraum", "Gespeichert", "Geändert", "Anmerkung")
        self.history_tree = ttk.Treeview(overview_frame, columns=hist_columns, show="headings",
                                          height=8, style="Modern.Treeview")
        self.history_tree.heading("Zeitraum", text="Zeitraum")
        self.history_tree.heading("Gespeichert", text="Gespeichert am")
        self.history_tree.heading("Geändert", text="Geändert")
        self.history_tree.heading("Anmerkung", text="Anmerkung")
        self.history_tree.column("Zeitraum", width=200, anchor="center")
        self.history_tree.column("Gespeichert", width=140, anchor="center")
        self.history_tree.column("Geändert", width=70, anchor="center")
        self.history_tree.column("Anmerkung", width=300, anchor="w")

        scrollbar_hist = ttk.Scrollbar(overview_frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar_hist.set)
        self.history_tree.pack(side="left", fill="both", expand=True)
        scrollbar_hist.pack(side="right", fill="y")

        # Tag-Style für manuell geänderte Pläne (gelber Hintergrund)
        self.history_tree.tag_configure("manuell", background="#FFF3B0")

        # Buttons
        btn_frame = ttk.Frame(history_frame)
        btn_frame.pack(fill="x", pady=(0, 10))
        ttk.Button(btn_frame, text="Plan anzeigen", command=self._view_plan_detail).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="Anmerkung bearbeiten", command=self._edit_plan_note).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="Plan löschen", command=self._delete_single_plan).pack(side="left", padx=2)

        # === UNTERER BEREICH: Detailansicht ===
        detail_frame = ttk.LabelFrame(history_frame, text=" Detailansicht ", padding=10)
        detail_frame.pack(fill="both", expand=True)

        detail_columns = ("Datum", "Tag", "VM", "NM", "Support")
        self.history_detail_tree = ttk.Treeview(detail_frame, columns=detail_columns, show="headings",
                                                 height=14, style="Modern.Treeview")
        for col_name, width in [("Datum", 85), ("Tag", 80), ("VM", 80), ("NM", 80), ("Support", 80)]:
            self.history_detail_tree.heading(col_name, text=col_name)
            self.history_detail_tree.column(col_name, width=width, minwidth=50, anchor="center")

        scrollbar_detail = ttk.Scrollbar(detail_frame, orient="vertical", command=self.history_detail_tree.yview)
        self.history_detail_tree.configure(yscrollcommand=scrollbar_detail.set)
        self.history_detail_tree.pack(side="left", fill="both", expand=True)
        scrollbar_detail.pack(side="right", fill="y")

        # Doppelklick zum Bearbeiten von Einträgen in der Detailansicht
        self.history_detail_tree.bind("<Double-1>", self._on_history_detail_double_click)
        self._hist_edit_entry = None
        self._hist_current_plan_idx = None  # Index des aktuell angezeigten Plans

        # Initial aktualisieren
        self.root.after(300, self.update_history_display)

    def update_history_display(self) -> None:
        """Aktualisiert die Übersicht der gespeicherten Pläne"""
        if not hasattr(self, 'history_tree'):
            return
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)

        # Pläne chronologisch sortiert anzeigen (neueste zuerst)
        sorted_plans = sorted(
            enumerate(self.saved_plans),
            key=lambda x: datetime.strptime(x[1].get("start_date", "01.01.1900"), "%d.%m.%Y"),
            reverse=True
        )

        for orig_idx, plan in sorted_plans:
            zeitraum = f"{plan.get('start_date', '?')} – {plan.get('end_date', '?')}"
            gespeichert = plan.get("saved_at", "?")
            is_manual = plan.get("manuell_geaendert", False)
            geaendert = "✓" if is_manual else "✗"
            anmerkung = plan.get("anmerkung", "")
            # Speichere den Original-Index als Tag für spätere Referenz; "manuell"-Tag färbt die Zeile
            row_tags = [str(orig_idx)]
            if is_manual:
                row_tags.append("manuell")
            self.history_tree.insert("", tk.END, values=(zeitraum, gespeichert, geaendert, anmerkung),
                                      tags=tuple(row_tags))

    def _get_selected_plan_index(self) -> Optional[int]:
        """Gibt den Index des ausgewählten Plans zurück"""
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("Warnung", "Bitte einen Plan auswählen!")
            return None
        tags = self.history_tree.item(selected[0], "tags")
        if tags:
            return int(tags[0])
        return None

    def _view_plan_detail(self) -> None:
        """Zeigt die Einträge des ausgewählten Plans in der Detailansicht"""
        plan_idx = self._get_selected_plan_index()
        if plan_idx is None:
            return

        self._hist_current_plan_idx = plan_idx
        plan = self.saved_plans[plan_idx]

        # Detail-Treeview leeren
        for item in self.history_detail_tree.get_children():
            self.history_detail_tree.delete(item)

        # Einträge anzeigen
        for entry in plan.get("entries", []):
            tag_kurz = entry.get("Wochentag", "")[:2] if entry.get("Wochentag") else ""
            self.history_detail_tree.insert(
                "", tk.END,
                values=(entry.get("Datum", ""), tag_kurz,
                       entry.get("Vormittag", ""), entry.get("Nachmittag", ""),
                       entry.get("Support", ""))
            )

    def _edit_plan_note(self) -> None:
        """Öffnet ein Dialogfenster zur Bearbeitung der Anmerkung"""
        plan_idx = self._get_selected_plan_index()
        if plan_idx is None:
            return

        plan = self.saved_plans[plan_idx]
        current_note = plan.get("anmerkung", "")

        # Dialog erstellen
        dialog = tk.Toplevel(self.root)
        dialog.title("Anmerkung bearbeiten")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Anmerkung:", font=('Segoe UI', 10, 'bold')).pack(padx=10, pady=(10, 5), anchor="w")

        text_widget = tk.Text(dialog, wrap=tk.WORD, width=45, height=5, font=('Segoe UI', 10))
        text_widget.pack(padx=10, pady=5, fill="both", expand=True)
        text_widget.insert("1.0", current_note)
        text_widget.focus_set()

        def save_note():
            new_note = text_widget.get("1.0", tk.END).strip()
            self.saved_plans[plan_idx]["anmerkung"] = new_note
            self.save_history()
            self.update_history_display()
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill="x", padx=10, pady=10)
        ttk.Button(btn_frame, text="Speichern", command=save_note).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Abbrechen", command=dialog.destroy).pack(side="left", padx=5)

        dialog.bind("<Escape>", lambda e: dialog.destroy())

    def _delete_single_plan(self) -> None:
        """Löscht einen einzelnen ausgewählten Plan"""
        plan_idx = self._get_selected_plan_index()
        if plan_idx is None:
            return

        plan = self.saved_plans[plan_idx]
        zeitraum = f"{plan.get('start_date', '?')} – {plan.get('end_date', '?')}"

        if messagebox.askyesno("Bestätigung", f"Plan '{zeitraum}' wirklich löschen?"):
            self.saved_plans.pop(plan_idx)
            self._hist_current_plan_idx = None
            # Detail-Treeview leeren
            for item in self.history_detail_tree.get_children():
                self.history_detail_tree.delete(item)
            self.save_history()
            self.update_history_display()
            self.update_evaluation_display()
            messagebox.showinfo("Erfolg", f"Plan '{zeitraum}' gelöscht!")

    def _on_history_detail_double_click(self, event) -> None:
        """Handler für Doppelklick zum Bearbeiten einer Zelle in der Detailansicht"""
        if self._hist_current_plan_idx is None:
            return

        if self._hist_edit_entry:
            self._hist_confirm_edit()

        region = self.history_detail_tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = self.history_detail_tree.identify_column(event.x)
        item = self.history_detail_tree.identify_row(event.y)

        if not item or not column:
            return

        col_idx = int(column.replace("#", ""))
        # Nur VM, NM, Support editierbar (Spalten 3, 4, 5)
        if col_idx < 3:
            return

        values = list(self.history_detail_tree.item(item, "values"))
        current_value = values[col_idx - 1] if col_idx - 1 < len(values) else ""

        bbox = self.history_detail_tree.bbox(item, column)
        if not bbox:
            return

        x, y, width, height = bbox

        self._hist_edit_entry = tk.Entry(self.history_detail_tree, width=width // 8)
        self._hist_edit_entry.place(x=x, y=y, width=width, height=height)
        self._hist_edit_entry.insert(0, current_value)
        self._hist_edit_entry.select_range(0, tk.END)
        self._hist_edit_entry.focus_set()

        self._hist_edit_item = item
        self._hist_edit_col_idx = col_idx
        self._hist_edit_row_idx = self.history_detail_tree.index(item)

        self._hist_edit_entry.bind("<Return>", lambda e: self._hist_confirm_edit())
        self._hist_edit_entry.bind("<Escape>", lambda e: self._hist_cancel_edit())
        self._hist_edit_entry.bind("<FocusOut>", lambda e: self._hist_confirm_edit())

    def _hist_confirm_edit(self) -> None:
        """Bestätigt die Bearbeitung in der Detailansicht und aktualisiert den gespeicherten Plan"""
        if not self._hist_edit_entry or self._hist_current_plan_idx is None:
            return

        new_value = self._hist_edit_entry.get().strip()

        # Treeview aktualisieren
        values = list(self.history_detail_tree.item(self._hist_edit_item, "values"))
        values[self._hist_edit_col_idx - 1] = new_value
        self.history_detail_tree.item(self._hist_edit_item, values=values)

        # Gespeicherten Plan aktualisieren
        col_map = {3: "Vormittag", 4: "Nachmittag", 5: "Support"}
        if self._hist_edit_col_idx in col_map:
            plan = self.saved_plans[self._hist_current_plan_idx]
            entries = plan.get("entries", [])
            if self._hist_edit_row_idx < len(entries):
                entries[self._hist_edit_row_idx][col_map[self._hist_edit_col_idx]] = new_value
                plan["manuell_geaendert"] = True
                self.save_history()
                self.update_history_display()

        self._hist_cleanup_edit()

    def _hist_cancel_edit(self) -> None:
        """Bricht die Bearbeitung in der Detailansicht ab"""
        self._hist_cleanup_edit()

    def _hist_cleanup_edit(self) -> None:
        """Räumt das Edit-Widget in der Detailansicht auf"""
        if self._hist_edit_entry:
            self._hist_edit_entry.destroy()
            self._hist_edit_entry = None
        self._hist_edit_item = None
        self._hist_edit_col_idx = None
        self._hist_edit_row_idx = None

    def create_help_tab(self, notebook: ttk.Notebook) -> None:
        """Erstellt Tab für Hilfe und Anleitung"""
        help_frame = ttk.Frame(notebook, padding=10)
        notebook.add(help_frame, text="Hilfe")

        text_area = scrolledtext.ScrolledText(help_frame, wrap=tk.WORD, width=100, height=30,
                                              font=('Segoe UI', 10))
        text_area.pack(fill="both", expand=True)

        # Tags für Formatierung
        text_area.tag_config("h1", font=('Segoe UI', 14, 'bold'), foreground="#2c3e50", spacing3=10)
        text_area.tag_config("h2", font=('Segoe UI', 12, 'bold'), foreground="#34495e", spacing3=5)
        text_area.tag_config("bold", font=('Segoe UI', 10, 'bold'))
        text_area.tag_config("italic", font=('Segoe UI', 10, 'italic'))

        help_content = [
            ("Regelwerk & Anleitung — Notdienst Schichtplaner", "h1"),
            ("\nDieses Regelwerk gilt für ALLE Mitarbeiter gleich. Jede Einteilung lässt sich damit "
             "nachvollziehen: Erst werden die harten Regeln (Abschnitt 2) angewendet, dann wählt die "
             "faire Auswahl (Abschnitt 3) unter den verbleibenden Verfügbaren. Verstöße — auch durch "
             "manuelle Änderungen — markiert die Plan-Prüfung automatisch rot.\n", ""),

            ("\n1. Mitarbeiter-Pools", "h2"),
            ("• ", ""), ("Pool A (Vormittag):", "bold"), (" kann den Vormittags-Notdienst alleine machen.\n"
             "• ", ""), ("Pool B (Vormittag mit Support):", "bold"), (" macht Vormittag nur zusammen mit einem Support.\n"
             "• ", ""), ("Pool C (Support):", "bold"), (" unterstützt Pool B am Vormittag.\n"
             "• ", ""), ("Pool D (Nachmittag):", "bold"), (" kann den Nachmittags-Notdienst machen.\n"
             "• ", ""), ("Pool E:", "bold"), (" ist freitags grundsätzlich nicht verfügbar.\n"
             "• ", ""), ("Pool F:", "bold"), (" ist Montag und Mittwoch grundsätzlich nicht verfügbar "
             "und hat maximal 1 Einsatz pro Woche.\n", ""),
            ("Hinweis: Im Fair-Modus hat die Reihenfolge innerhalb der Pools KEINEN Einfluss — "
             "es zählen nur die Einsatzzahlen. (Nur im klassischen Modus bestimmt sie die Rotation.)", "italic"),

            ("\n2. Harte Regeln (werden bei der Planung nie verletzt)", "h2"),
            ("R1 — Abwesenheiten:", "bold"),
            (" Wer eingetragen ist, wird nicht eingeteilt. 'Ganztags' = keine Schicht; "
             "'Nur vormittags' = kein Vormittag und kein Support, Nachmittag möglich; "
             "'Nur nachmittags' = kein Nachmittag, Vormittag/Support möglich.\n", ""),
            ("R2 — Vortags-Regel:", "bold"),
            (" Wer am nächsten Arbeitstag ganztags oder vormittags abwesend ist, bekommt am Tag davor "
             "keinen Nachmittagsdienst. Gilt auch Freitag → Montag.\n", ""),
            ("R3 — Rückkehr-Regel:", "bold"),
            (" Wer am vorherigen Arbeitstag ganztags abwesend war, wird am Folgetag nicht vormittags "
             "eingeteilt (kein Vormittag, kein Support). Gilt auch Freitag → Montag.\n", ""),
            ("R4 — Ruhezeit:", "bold"),
            (" Wer gestern Nachmittag hatte, bekommt heute keinen Vormittag. Gilt auch Freitag → Montag.\n", ""),
            ("R5 — Keine Doppelbelegung:", "bold"),
            (" Niemand hat zwei Schichten am selben Tag. (Anzeige-Ausnahme: Steht der Vormittags-MA auch "
             "in der Support-Spalte, heißt das nur 'kein separater Support nötig' — das ist KEIN zweiter Einsatz.)\n", ""),
            ("R6 — Support-Pflicht:", "bold"),
            (" Ein Pool-B-MA am Vormittag bekommt immer einen Support aus Pool C dazu.\n", ""),
            ("R7 — Wochentags-Sperren:", "bold"),
            (" Pool E wird freitags nie eingeteilt, Pool F montags und mittwochs nie.\n", ""),
            ("R8 — Pool-F-Limit:", "bold"),
            (" Wer in Pool F steht, hat maximal 1 Einsatz pro Woche.\n", ""),
            ("R9 — Feiertag:", "bold"),
            (" Der eingetragene Notdienst-MA übernimmt den kompletten Feiertag und ist dafür den Rest "
             "der Woche von regulären Diensten befreit.\n", ""),
            ("R10 — Montag Woche 2:", "bold"),
            (" Am Montag der zweiten Woche wird wegen höheren Aufkommens immer ein separater Support "
             "eingeplant (abschaltbar über den Haken 'Montag = normaler Tag').\n", ""),

            ("\n3. Faire Auswahl — wer kommt dran?", "h2"),
            ("Unter allen MA, die nach den harten Regeln verfügbar sind, entscheiden diese Kriterien "
             "in genau dieser Reihenfolge (bei Gleichstand zählt das nächste):\n", ""),
            ("1. Wenigste Einsätze in der LAUFENDEN WOCHE", "bold"),
            (" — niemand bekommt z.B. den 3. Einsatz, solange ein Verfügbarer erst 1 hat.\n", ""),
            ("2. Wenigste Einsätze dieses Schichttyps", "bold"),
            (" (VM, NM oder Support — inkl. Historie).\n", ""),
            ("3. Geringste Gesamtlast", "bold"),
            (" (VM + NM + Support zusammen, inkl. Historie laut eingestelltem Horizont).\n", ""),
            ("4. Gestern nicht eingeteilt gewesen.\n", ""),
            ("5. Am längsten nicht denselben Schichttyp gehabt.\n", ""),
            ("6. Am längsten gar keinen Einsatz gehabt.\n", ""),
            ("Die Einsatz-Zähler fließen dabei nach Rollen-Einsetzbarkeit gewichtet in den Vergleich ein, "
             "damit die Belastung über alle Aufgabenarten hinweg ausgewogen bleibt.\n", "italic"),
            ("\nZusätzlich gilt: Derselbe Schichttyp frühestens nach 3 Arbeitstagen wieder — diese Pause "
             "wird nur gelockert, wenn sonst niemand verfügbar wäre.\n", ""),
            ("Optimierungspass:", "bold"),
            (" Nach der Berechnung prüft das Programm automatisch Tausch- und Ersetzungsmöglichkeiten und "
             "übernimmt nur Änderungen, die die Verteilung messbar gleichmäßiger machen — ohne eine harte "
             "Regel zu verletzen.\n", ""),
            ("Historie-Horizont:", "bold"),
            (" Wie weit die Vergangenheit in die Zähler einfließt, ist einstellbar (Standard: 3 Monate). "
             "So gleichen sich Unterschiede aus früheren Wochen automatisch wieder aus.\n", ""),

            ("\n4. Zählweise (so wird gezählt)", "h2"),
            ("• Jeder Vormittag, Nachmittag und Support-Einsatz zählt 1.\n"
             "• Steht der Vormittags-MA zugleich in der Support-Spalte ('kein separater Support nötig'), "
             "zählt das NICHT doppelt — es bleibt 1 Einsatz.\n"
             "• Ein Feiertags-Notdienst zählt als 2 Einsätze (Vormittag + Nachmittag), dafür ist der MA "
             "den Rest der Woche befreit.\n"
             "• Die Tabelle 'Wochenlast' (rechts unten) zeigt die Einsätze pro MA für Woche 1, Woche 2 "
             "und gesamt; der Tab 'Auswertung' zeigt die Langzeit-Statistik.\n", ""),

            ("\n5. Planung erstellen (Bedienung)", "h2"),
            ("1. Startdatum (Montag) ist automatisch vorbefüllt — nächster Montag nach dem letzten Plan.\n"
             "2. ", ""), ("Faire Verteilung (Standard):", "bold"), (" Die Tag-1-Felder können leer bleiben, "
             "Tag 1 wird fair mitgeplant. Eine manuelle Vorgabe ist möglich.\n"
             "   ", ""), ("Klassischer Modus:", "bold"), (" Tag 1 muss als Startpunkt der Rotation angegeben werden.\n"
             "3. 'Planung erstellen' klicken — berechnet werden 10 Arbeitstage (2 Wochen, Mo–Fr).\n"
             "4. Passt alles, den Plan mit 'Plan speichern' in die Historie übernehmen — erst dann fließt "
             "er in Statistik und künftige Fairness-Berechnung ein.\n", ""),

            ("\n6. Abwesenheiten / Termine erfassen", "h2"),
            ("Rechts im Planungs-Tab: Kürzel (mehrere mit Komma), Von/Bis-Datum (Bis leer = nur ein Tag) "
             "und Zeitraum (Ganztags / Nur vormittags / Nur nachmittags) wählen, dann '+ Hinzufügen'.\n"
             "Abwesenheiten werden automatisch gespeichert und sind nach einem Neustart wieder da.\n"
             "Unbekannte Kürzel meldet das Programm sofort (Tippfehler-Schutz).\n", ""),

            ("\n7. Plan-Prüfung (automatische Kontrolle)", "h2"),
            ("Unter dem Planungsergebnis prüft das Programm jeden Plan automatisch gegen ALLE Regeln aus "
             "Abschnitt 2 — auch nach jeder manuellen Änderung per Doppelklick:\n"
             "• ", ""), ("Rote Zeilen = Regelverstoß", "bold"), (" (z.B. abwesender MA eingeteilt, Pool B ohne "
             "Support, Ruhezeit verletzt, Pool-F-Limit überschritten, Tag unbesetzt).\n"
             "• ", ""), ("Orange Zeilen = Hinweis", "bold"), (" (z.B. unbekanntes Kürzel, Montag Woche 2 ohne "
             "separaten Support, derselbe Schichttyp in kurzem Abstand wiederholt, auffällig ungleiche "
             "Wochenverteilung).\n"
             "Die Liste darunter benennt jeden Befund konkret mit Datum und Name. Damit ist jede "
             "Einteilung transparent überprüfbar — für den Planer und für das Team.\n", ""),

            ("\n8. Nachhol-System (nur klassischer Modus)", "h2"),
            ("Im klassischen Rotations-Modus wird ein übersprungener MA in eine Nachhol-Queue aufgenommen: "
             "Nach 2 Tagen wird er bevorzugt, nach 4 Tagen mit absoluter Priorität eingeplant. "
             "Im Fair-Modus ist das nicht nötig — dort gleichen die Einsatz-Zähler Übersprünge "
             "automatisch aus.\n", "")
        ]

        for text, tag in help_content:
            text_area.insert(tk.END, text, tag)
        
        text_area.configure(state='disabled')  # Schreibgeschützt

    def save_current_plan(self) -> None:
        """Speichert den aktuellen Plan in die Historie"""
        if not self.planning_result:
            messagebox.showwarning("Warnung", "Keine Planung vorhanden!")
            return

        # Datum des Plans ermitteln
        start_date = self.planning_result[0]["Datum"] if self.planning_result else ""
        end_date = self.planning_result[-1]["Datum"] if self.planning_result else ""

        # Prüfen ob Plan bereits existiert
        for plan in self.saved_plans:
            if plan.get("start_date") == start_date and plan.get("end_date") == end_date:
                if messagebox.askyesno("Plan existiert", "Ein Plan für diesen Zeitraum existiert bereits. Überschreiben?"):
                    self.saved_plans.remove(plan)
                    break
                else:
                    return

        # Plan speichern
        new_plan = {
            "start_date": start_date,
            "end_date": end_date,
            "saved_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "entries": [row.copy() for row in self.planning_result],
            "anmerkung": "",
            "manuell_geaendert": self._plan_manually_edited
        }
        self.saved_plans.append(new_plan)
        self._plan_manually_edited = False  # Reset nach Speichern
        self.save_history()
        messagebox.showinfo("Erfolg", f"Plan vom {start_date} bis {end_date} gespeichert!")
        self.update_evaluation_display()
        self.update_history_display()

    def update_evaluation_display(self) -> None:
        """Aktualisiert die Statistik-Anzeige basierend auf Filter"""
        # Filter lesen
        filter_from = self.eval_from_var.get().strip() if hasattr(self, 'eval_from_var') else ""
        filter_to = self.eval_to_var.get().strip() if hasattr(self, 'eval_to_var') else ""

        # Statistik berechnen
        stats = self._calculate_statistics(filter_from, filter_to)

        # Treeview leeren
        for item in self.stats_tree.get_children():
            self.stats_tree.delete(item)

        # Statistik sortieren: Nach Gesamt (absteigend), dann Name (aufsteigend)
        # Wir erstellen erst eine Liste mit (Name, Counts, Total)
        stats_list = []
        for ma, counts in stats.items():
            total = counts["VM"] + counts["NM"] + counts["Support"]
            stats_list.append((ma, counts, total))

        # Sortieren: -total für absteigend, name für aufsteigend
        stats_list.sort(key=lambda x: (-x[2], x[0]))

        # Statistik anzeigen
        for ma, counts, total in stats_list:
            self.stats_tree.insert("", tk.END, values=(
                ma, counts["VM"], counts["NM"], counts["Support"], total
            ))

        # Info aktualisieren — inkl. Fairness-Metrik (σ + Spread)
        plan_count = len(self.saved_plans)
        metrics = self._compute_fairness_metrics(stats)
        metric_text = ""
        if metrics:
            metric_text = (
                f"  ·  σ={metrics['sigma']:.2f}  "
                f"Min: {metrics['min_ma']} ({metrics['min_value']})  "
                f"Max: {metrics['max_ma']} ({metrics['max_value']})  "
                f"Spread: {metrics['spread']}"
            )
        if plan_count > 0:
            date_range = ""
            if filter_from or filter_to:
                date_range = f" (Filter: {filter_from or '...'} bis {filter_to or '...'})"
            self.stats_info_label.config(text=f"{plan_count} Plan(e) gespeichert.{date_range}{metric_text}")
        else:
            self.stats_info_label.config(text="Keine Pläne gespeichert.")

    def _calculate_statistics(self, filter_from: str, filter_to: str) -> Dict[str, Dict[str, int]]:
        """Berechnet Statistik über alle gespeicherten Pläne"""
        stats: Dict[str, Dict[str, int]] = {}

        # Filter-Daten parsen
        from_date = None
        to_date = None
        try:
            if filter_from:
                from_date = datetime.strptime(filter_from, "%d.%m.%Y")
            if filter_to:
                to_date = datetime.strptime(filter_to, "%d.%m.%Y")
        except ValueError:
            pass  # Ungültige Filter ignorieren

        for plan in self.saved_plans:
            for entry in plan.get("entries", []):
                # Datum prüfen
                try:
                    entry_date = datetime.strptime(entry["Datum"], "%d.%m.%Y")
                    if from_date and entry_date < from_date:
                        continue
                    if to_date and entry_date > to_date:
                        continue
                except ValueError:
                    continue

                # Zählen
                # WICHTIG: Wenn Support == Vormittag, ist das nur die Anzeige-Duplizierung
                # ("kein separater Support nötig") und zählt NICHT als eigener Einsatz.
                vm_ma = entry.get("Vormittag", "").strip()
                for shift_type, key in [("Vormittag", "VM"), ("Nachmittag", "NM"), ("Support", "Support")]:
                    ma = entry.get(shift_type, "").strip()
                    if not ma:
                        continue
                    if key == "Support" and ma == vm_ma:
                        continue
                    if ma not in stats:
                        stats[ma] = {"VM": 0, "NM": 0, "Support": 0}
                    stats[ma][key] += 1

        # Manuelle Korrekturen anwenden
        for ma, corrections in self.manual_stats_corrections.items():
            if ma not in stats:
                 # Wenn MA nicht in den Plänen, aber eine Korrektur existiert, trotzdem anzeigen
                 if any(corrections.values()): # Nur wenn Korrektur != 0
                     stats[ma] = {"VM": 0, "NM": 0, "Support": 0}
            
            if ma in stats:
                for key in ["VM", "NM", "Support"]:
                    stats[ma][key] += corrections.get(key, 0)

        return stats

    def _save_filter(self) -> None:
        """Speichert den aktuellen Datumsfilter"""
        self._eval_filter_from = self.eval_from_var.get().strip()
        self._eval_filter_to = self.eval_to_var.get().strip()
        self.save_history()
        messagebox.showinfo("Erfolg", "Filter gespeichert!")

    def _on_stats_double_click(self, event) -> None:
        """Handler für Doppleclick auf Statistik"""
        # Aktives Eingabefeld schließen
        if getattr(self, '_stats_edit_entry', None):
            self._stats_cancel_edit()

        region = self.stats_tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = self.stats_tree.identify_column(event.x)
        item = self.stats_tree.identify_row(event.y)

        if not item or not column:
            return

        # Spaltenindex: #1=MA, #2=VM, #3=NM, #4=Support, #5=Gesamt
        col_idx = int(column.replace("#", ""))
        
        # MA und Gesamt nicht editierbar
        if col_idx == 1 or col_idx == 5:
            return

        # Aktuelle Werte
        values = list(self.stats_tree.item(item, "values"))
        current_val = values[col_idx - 1]

        bbox = self.stats_tree.bbox(item, column)
        if not bbox:
            return
        
        x, y, width, height = bbox

        self._stats_edit_entry = tk.Entry(self.stats_tree, width=width // 8)
        self._stats_edit_entry.place(x=x, y=y, width=width, height=height)
        self._stats_edit_entry.insert(0, current_val)
        self._stats_edit_entry.select_range(0, tk.END)
        self._stats_edit_entry.focus_set()

        # Metadaten speichern
        self._stats_edit_item = item
        self._stats_edit_col_idx = col_idx
        # MA Name für Speicherung
        self._stats_edit_ma = values[0] 

        self._stats_edit_entry.bind("<Return>", lambda e: self._confirm_stats_edit())
        self._stats_edit_entry.bind("<Escape>", lambda e: self._stats_cancel_edit())
        self._stats_edit_entry.bind("<FocusOut>", lambda e: self._confirm_stats_edit())

    def _confirm_stats_edit(self) -> None:
        """Bestätigt Statistik-Bearbeitung"""
        if not getattr(self, '_stats_edit_entry', None):
            return

        try:
            new_val_str = self._stats_edit_entry.get().strip()
            new_val = int(new_val_str)
        except ValueError:
            # Ungültige Eingabe -> Abbruch
            self._stats_cancel_edit()
            return
        
        # Wir müssen den Delta berechnen. 
        # Das Problem: Der angezeigte Wert ist (Berechnet + Manuell).
        # Wir wollen Manuell_Neu so setzen, dass (Berechnet + Manuell_Neu) = Eingabe_Neu
        # => Manuell_Neu = Eingabe_Neu - Berechnet
        
        # 1. Berechneten Wert (ohne Korrektur) neu ermitteln
        filter_from = self.eval_from_var.get().strip() if hasattr(self, 'eval_from_var') else ""
        filter_to = self.eval_to_var.get().strip() if hasattr(self, 'eval_to_var') else ""
        
        # Temporär Korrekturen deaktivieren um Basiswert zu bekommen
        current_corrections = self.manual_stats_corrections
        self.manual_stats_corrections = {} 
        base_stats = self._calculate_statistics(filter_from, filter_to)
        self.manual_stats_corrections = current_corrections # Restore

        ma = self._stats_edit_ma
        col_map = {2: "VM", 3: "NM", 4: "Support"}
        key = col_map[self._stats_edit_col_idx]

        base_val = 0
        if ma in base_stats:
            base_val = base_stats[ma].get(key, 0)
        
        # Neue Korrektur berechnen
        correction = new_val - base_val

        # Speichern
        if ma not in self.manual_stats_corrections:
            self.manual_stats_corrections[ma] = {}
        
        self.manual_stats_corrections[ma][key] = correction
        
        # Wenn Korrektur 0 ist, Eintrag ggf. bereinigen
        if correction == 0:
            if key in self.manual_stats_corrections[ma]:
                del self.manual_stats_corrections[ma][key]
            if not self.manual_stats_corrections[ma]:
                del self.manual_stats_corrections[ma]

        self.save_history()
        self._stats_cancel_edit()
        self.update_evaluation_display()

    def _stats_cancel_edit(self) -> None:
        if getattr(self, '_stats_edit_entry', None):
            self._stats_edit_entry.destroy()
            self._stats_edit_entry = None

    def _clear_all_plans(self) -> None:
        """Löscht alle gespeicherten Pläne"""
        if not self.saved_plans:
            messagebox.showinfo("Info", "Keine Pläne vorhanden.")
            return

        if messagebox.askyesno("Bestätigung", f"{len(self.saved_plans)} Plan(e) wirklich löschen?"):
            self.saved_plans = []
            self.save_history()
            self.update_evaluation_display()
            self.update_history_display()
            messagebox.showinfo("Erfolg", "Alle Pläne gelöscht!")

    # -------------------- Abwesenheiten --------------------

    def update_employee_list(self) -> None:
        """Aktualisiert die Mitarbeiterliste in der Combobox"""
        self._cache_dirty = True
        all_employees = self._get_all_employees()
        self.employee_combo['values'] = all_employees
        messagebox.showinfo("Info", f"Mitarbeiterliste aktualisiert: {len(all_employees)} Mitarbeiter gefunden")

    def add_absence(self) -> None:
        """Fügt Abwesenheiten hinzu: Datumsbereich + Zeitraum (ganztags/nur VM/nur NM).

        Unterstützt mehrere komma-getrennte Kürzel. 'Bis' leer = nur der 'Von'-Tag.
        """
        employee = self.employee_var.get().strip()
        from_str = self.absence_from_var.get().strip()
        to_str = self.absence_to_var.get().strip()
        scope_label = self.absence_scope_var.get().strip()

        if not employee or not from_str:
            messagebox.showwarning("Warnung", "Bitte Mitarbeiter und 'Von'-Datum eingeben!")
            return

        try:
            from_date = datetime.strptime(from_str, "%d.%m.%Y")
            to_date = datetime.strptime(to_str, "%d.%m.%Y") if to_str else from_date
        except ValueError:
            messagebox.showerror("Fehler", "Ungültiges Datum! Format: TT.MM.JJJJ (z.B. 12.01.2026)")
            return

        if to_date < from_date:
            messagebox.showerror("Fehler", "'Bis'-Datum liegt vor 'Von'-Datum!")
            return

        scope = "ganz"
        for label, key in ABSENCE_SCOPE_OPTIONS:
            if label == scope_label:
                scope = key
                break

        employees = self._parse_employee_input(employee)
        if not employees:
            messagebox.showwarning("Warnung", "Keine gültigen Mitarbeiter eingegeben!")
            return

        # Kürzel-Validierung: unbekannte Kürzel nur nach Rückfrage übernehmen
        known = set(self._get_all_employees())
        unknown = [e for e in employees if e not in known]
        if unknown:
            if not messagebox.askyesno(
                "Unbekanntes Kürzel",
                f"Folgende Kürzel sind in keinem Pool eingetragen: {', '.join(unknown)}\n"
                "Trotzdem hinzufügen?"
            ):
                return

        added_count = 0
        for emp in employees:
            new_entry = {"ma": emp, "von": from_date, "bis": to_date, "zeit": scope}
            duplicate = any(
                a["ma"] == emp and a["von"] == from_date and a["bis"] == to_date and a["zeit"] == scope
                for a in self.absences
            )
            if not duplicate:
                self.absences.append(new_entry)
                added_count += 1

        if added_count > 0:
            self.update_absence_display()
            self._save_absences()
            # Eingabefelder für schnelle Mehrfacheingabe vorbereiten (MA bleibt stehen)
            self.absence_from_var.set("")
            self.absence_to_var.set("")

    def remove_absence(self) -> None:
        """Entfernt ausgewählte Abwesenheiten"""
        selected = self.absence_tree.selection()
        if not selected:
            messagebox.showwarning("Warnung", "Bitte einen Eintrag auswählen!")
            return

        # Original-Indizes aus den Tags lesen und absteigend löschen
        indices = []
        for item in selected:
            tags = self.absence_tree.item(item, "tags")
            if tags:
                try:
                    indices.append(int(tags[0]))
                except (ValueError, IndexError):
                    continue

        for idx in sorted(set(indices), reverse=True):
            if 0 <= idx < len(self.absences):
                self.absences.pop(idx)

        self.update_absence_display()
        self._save_absences()

    def update_absence_display(self) -> None:
        """Aktualisiert die Anzeige der Abwesenheiten"""
        for item in self.absence_tree.get_children():
            self.absence_tree.delete(item)

        # Sortiert nach Von-Datum, dann MA; Original-Index als Tag für das Entfernen
        sorted_entries = sorted(enumerate(self.absences), key=lambda x: (x[1]["von"], x[1]["ma"]))
        for orig_idx, entry in sorted_entries:
            von_str = entry["von"].strftime("%d.%m.%Y")
            bis_str = entry["bis"].strftime("%d.%m.%Y")
            if bis_str == von_str:
                bis_str = ""
            zeit_label = ABSENCE_SCOPE_LABELS.get(entry["zeit"], entry["zeit"])
            self.absence_tree.insert("", tk.END,
                                     values=(entry["ma"], von_str, bis_str, zeit_label),
                                     tags=(str(orig_idx),))

    def _load_absences(self) -> None:
        """Lädt gespeicherte Abwesenheiten aus eigener Datei (überleben Neustarts)."""
        self.absences = []
        try:
            if os.path.exists(ABSENCES_FILE):
                with open(ABSENCES_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                for entry in data.get("absences", []):
                    try:
                        self.absences.append({
                            "ma": entry["ma"],
                            "von": datetime.strptime(entry["von"], "%d.%m.%Y"),
                            "bis": datetime.strptime(entry["bis"], "%d.%m.%Y"),
                            "zeit": entry.get("zeit", "ganz"),
                        })
                    except (KeyError, ValueError):
                        continue
        except Exception:
            self.absences = []

    def _save_absences(self) -> None:
        """Speichert Abwesenheiten still in eigene Datei (kein Dialog)."""
        try:
            data = {"absences": [
                {"ma": a["ma"],
                 "von": a["von"].strftime("%d.%m.%Y"),
                 "bis": a["bis"].strftime("%d.%m.%Y"),
                 "zeit": a["zeit"]}
                for a in self.absences
            ]}
            with open(ABSENCES_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def _absence_sets_for_date(self, date: datetime) -> Tuple[set, set]:
        """Gibt (vormittags_abwesend, nachmittags_abwesend) für ein Datum zurück.

        'vm'-Abwesenheit blockiert Vormittag UND Support, 'nm' nur Nachmittag,
        'ganz' blockiert alles.
        """
        absent_vm = set()
        absent_nm = set()
        day = date.replace(hour=0, minute=0, second=0, microsecond=0)
        for entry in self.absences:
            if entry["von"] <= day <= entry["bis"]:
                if entry["zeit"] in ("ganz", "vm"):
                    absent_vm.add(entry["ma"])
                if entry["zeit"] in ("ganz", "nm"):
                    absent_nm.add(entry["ma"])
        return absent_vm, absent_nm

    def _full_day_absent_on(self, date: datetime) -> set:
        """MA, die an einem Datum GANZTAGS abwesend eingetragen sind.

        Nur echte Abwesenheits-Einträge — Pool-E/F-Regeln und Feiertage zählen
        hier bewusst nicht (sonst würden Dauer-Regelungen wie Pool F jede Woche
        die Folgetags-Logik auslösen).
        """
        day = date.replace(hour=0, minute=0, second=0, microsecond=0)
        return {
            a["ma"] for a in self.absences
            if a["zeit"] == "ganz" and a["von"] <= day <= a["bis"]
        }

    def _prefill_start_date(self) -> None:
        """Befüllt das Startdatum automatisch: nächster Montag nach dem letzten Plan,
        sonst der nächste kommende Montag. Nur wenn das Feld leer ist."""
        try:
            if self.start_date_entry.get().strip():
                return
            tag10_data = self._get_last_tag10_data()
            if tag10_data:
                start_str = tag10_data["start_date"]
            else:
                today = datetime.now()
                days_until_monday = (7 - today.weekday()) % 7
                if days_until_monday == 0:
                    days_until_monday = 7
                start_str = (today + timedelta(days=days_until_monday)).strftime("%d.%m.%Y")
            self.start_date_entry.insert(0, start_str)
            self._update_absence_date_options()
        except Exception:
            pass

    def _update_absence_date_options(self) -> None:
        """Befüllt die Von/Bis-Dropdowns mit den 10 Arbeitstagen des aktuellen Plans."""
        try:
            start_date = datetime.strptime(self.start_date_entry.get().strip(), "%d.%m.%Y")
        except (ValueError, AttributeError):
            return
        dates = []
        for day in self._iterate_working_days(start_date):
            dates.append(day["current_date"].strftime("%d.%m.%Y"))
        if hasattr(self, 'absence_from_combo'):
            self.absence_from_combo['values'] = dates
            self.absence_to_combo['values'] = dates

    # -------------------- Feiertage (dauerhaft gespeichert) --------------------

    def add_holiday(self) -> None:
        """Fügt einen Feiertag hinzu und speichert in Config"""
        date_str = self.holiday_date_var.get().strip()
        name = self.holiday_name_var.get().strip()
        employee = self.holiday_employee_var.get().strip()

        if not date_str or not name:
            messagebox.showwarning("Warnung", "Bitte Datum und Name eingeben!")
            return

        # Datum validieren (TT.MM Format)
        try:
            parts = date_str.split(".")
            if len(parts) != 2:
                raise ValueError("Format muss TT.MM sein")
            day, month = int(parts[0]), int(parts[1])
            if not (1 <= day <= 31 and 1 <= month <= 12):
                raise ValueError("Ungültiges Datum")
            date_str = f"{day:02d}.{month:02d}"  # Normalisieren
        except (ValueError, IndexError) as e:
            messagebox.showerror("Fehler", f"Ungültiges Datum: {e}\nFormat: TT.MM (z.B. 25.12)")
            return

        # Prüfen ob Datum bereits existiert
        for holiday in self.config["feiertage"]:
            if holiday["datum"] == date_str:
                # Update statt neu hinzufügen
                holiday["name"] = name
                holiday["mitarbeiter"] = employee
                self.update_holiday_display()
                self.save_config()
                self._clear_holiday_inputs()
                return

        # Neuen Feiertag hinzufügen
        self.config["feiertage"].append({
            "datum": date_str,
            "name": name,
            "mitarbeiter": employee
        })

        # Nach Datum sortieren
        self.config["feiertage"].sort(key=lambda x: (int(x["datum"].split(".")[1]), int(x["datum"].split(".")[0])))

        self.update_holiday_display()
        self.save_config()
        self._clear_holiday_inputs()

    def _clear_holiday_inputs(self) -> None:
        """Leert die Feiertags-Eingabefelder"""
        self.holiday_date_var.set("")
        self.holiday_name_var.set("")
        self.holiday_employee_var.set("")

    def remove_holiday(self) -> None:
        """Entfernt einen Feiertag und speichert Config"""
        selected = self.holiday_tree.selection()
        if not selected:
            messagebox.showwarning("Warnung", "Bitte einen Eintrag auswählen!")
            return

        for item in selected:
            values = self.holiday_tree.item(item, 'values')
            date_str = values[0]
            # Feiertag aus Liste entfernen
            self.config["feiertage"] = [
                h for h in self.config["feiertage"] if h["datum"] != date_str
            ]

        self.update_holiday_display()
        self.save_config()

    def update_holiday_display(self) -> None:
        """Aktualisiert die Anzeige der Feiertage aus Config"""
        for item in self.holiday_tree.get_children():
            self.holiday_tree.delete(item)

        for holiday in self.config.get("feiertage", []):
            self.holiday_tree.insert("", tk.END, values=(
                holiday["datum"],
                holiday["name"],
                holiday.get("mitarbeiter", "")
            ))

    def _get_holiday_employees_for_week(self, tag_nr: int, start_date: datetime) -> List[str]:
        """Gibt Mitarbeiter zurück, die in der Woche des Tags Feiertags-Notdienst haben"""
        # Woche berechnen (0-4 = Woche 1, 5-9 = Woche 2)
        week_number = tag_nr // 5  # 0 oder 1
        week_start = week_number * 5
        week_end = week_start + 5

        holiday_employees = []

        # Alle Arbeitstage der Woche durchgehen (Mo-Fr)
        for working_day in range(week_start, week_end):
            weekday_in_week = working_day % 5  # 0=Mo, 1=Di, 2=Mi, 3=Do, 4=Fr
            wk = working_day // 5
            current_date = start_date + timedelta(days=weekday_in_week + (wk * 7))
            date_str = current_date.strftime("%d.%m")

            # Prüfen ob dieser Tag ein Feiertag ist
            for holiday in self.config.get("feiertage", []):
                ma = holiday.get("mitarbeiter", "")
                if holiday["datum"] == date_str and ma and ma != '-- Kein Notdienst --':
                    holiday_employees.append(ma)

        return holiday_employees

    def save_pools(self) -> None:
        """Speichert die Pool-Konfiguration (optimiert mit DRY)"""
        try:
            pool_entries = {
                "pool_vm_alle": self.pool_vm_alle_entry,
                "pool_vm_teilweise": self.pool_vm_teilweise_entry,
                "pool_vm_support": self.pool_vm_support_entry,
                "pool_nm_alle": self.pool_nm_alle_entry,
                "pool_freitag_abwesend": self.pool_freitag_abwesend_entry,
                "pool_mo_mi_abwesend": self.pool_mo_mi_abwesend_entry
            }

            for pool_key, entry_widget in pool_entries.items():
                self.config[pool_key] = self._parse_employee_input(entry_widget.get())

            self.save_config()
            self._auto_update_employee_list()  # Automatisch Mitarbeiterliste aktualisieren
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Speichern der Pools: {e}")

    # -------------------- Planung --------------------

    def _validate_planning_inputs(self) -> Optional[Tuple[datetime, str, str, str]]:
        """Validiert Eingaben für Planung. Gibt (start_date, first_vm, first_nm, first_support) oder None zurück"""
        start_date_str = self.start_date_entry.get().strip()
        if not start_date_str:
            messagebox.showerror("Fehler", "Bitte Startdatum eingeben!")
            return None

        try:
            start_date = datetime.strptime(start_date_str, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Fehler", "Ungültiges Datumsformat! Bitte TT.MM.YYYY verwenden.")
            return None

        if start_date.weekday() != 0:  # Montag
            messagebox.showerror("Fehler", "Startdatum muss ein Montag sein!")
            return None

        first_vm = self.first_vm_entry.get().strip()
        first_nm = self.first_nm_entry.get().strip()
        first_support = self.first_support_entry.get().strip()

        # Im Fair-Modus ist Tag 1 optional (wird sonst fair mitgeplant);
        # im klassischen Modus weiterhin Pflicht (Startpunkt der Rotation).
        fair_mode = bool(self.fair_mode_var.get()) if hasattr(self, 'fair_mode_var') else False
        if not fair_mode and not all([first_vm, first_nm]):
            messagebox.showerror("Fehler", "Bitte mindestens Vormittag und Nachmittag für den ersten Tag ausfüllen!")
            return None

        # Kürzel-Validierung der Tag-1-Eingaben (Tippfehler-Schutz)
        known = set(self._get_all_employees())
        unknown = [x for x in (first_vm, first_nm, first_support) if x and x not in known]
        if unknown:
            if not messagebox.askyesno(
                "Unbekanntes Kürzel",
                f"Folgende Kürzel sind in keinem Pool eingetragen: {', '.join(sorted(set(unknown)))}\n"
                "Trotzdem fortfahren?"
            ):
                return None

        return start_date, first_vm, first_nm, first_support

    def _initialize_pool_positions(self, first_vm: str, first_nm: str, first_support: str) -> Dict[str, int]:
        """Initialisiert Pool-Positionen basierend auf ersten Mitarbeitern"""
        pool_positions = {"vm_alle": 0, "vm_teilweise": 0, "vm_support": 0, "nm_alle": 0}

        # Effizienter: nur einmal suchen
        pool_configs = [
            ("vm_alle", first_vm, self.config["pool_vm_alle"]),
            ("nm_alle", first_nm, self.config["pool_nm_alle"]),
            ("vm_support", first_support, self.config["pool_vm_support"]),
            ("vm_teilweise", first_vm, self.config["pool_vm_teilweise"])
        ]

        for pool_key, employee, pool_list in pool_configs:
            if employee and pool_list and employee in pool_list:
                idx = pool_list.index(employee)
                pool_positions[pool_key] = (idx + 1) % len(pool_list)

        return pool_positions

    def _get_absent_for_day(self, working_day: int, current_date: datetime, is_friday: bool,
                            is_monday: bool, is_wednesday: bool, start_date: datetime) -> Dict[str, List[str]]:
        """Gibt abwesende Mitarbeiter für einen Tag zurück, getrennt nach Tageshälfte.

        Rückgabe: {"VM": [...], "NM": [...]}
        Die VM-Liste gilt für Vormittag UND Support (beides Vormittags-Einsätze).
        """
        absent_vm, absent_nm = self._absence_sets_for_date(current_date)

        # Vortags-Regel: Wer am NÄCHSTEN Arbeitstag abwesend ist (ganztags oder
        # vormittags), bekommt am Tag davor keinen Nachmittagsdienst.
        # Freitag -> nächster Arbeitstag ist der Montag.
        next_work_date = current_date + timedelta(days=3 if current_date.weekday() == 4 else 1)
        next_day_vm_absent, _ = self._absence_sets_for_date(next_work_date)
        absent_nm.update(next_day_vm_absent)

        # Rückkehr-Regel: Wer am VORHERIGEN Arbeitstag ganztags abwesend war,
        # wird am Folgetag nicht vormittags eingeteilt (kein VM, kein Support).
        # Montag -> vorheriger Arbeitstag ist der Freitag.
        prev_work_date = current_date - timedelta(days=3 if current_date.weekday() == 0 else 1)
        absent_vm.update(self._full_day_absent_on(prev_work_date))

        # Ganztägige Pool-Regeln gelten für beide Tageshälften
        full_day = set()
        if is_friday:
            full_day.update(self.config.get("pool_freitag_abwesend", []))
        if is_monday or is_wednesday:
            full_day.update(self.config.get("pool_mo_mi_abwesend", []))
        # Feiertags-Notdienst-MA für die ganze Woche ausschließen
        full_day.update(self._get_holiday_employees_for_week(working_day, start_date))

        absent_vm.update(full_day)
        absent_nm.update(full_day)

        return {"VM": sorted(absent_vm), "NM": sorted(absent_nm)}

    def _find_employee_from_pool(self, pool: List[str], start_pos: int,
                                 excluded: List[str], absent: List[str]) -> Optional[Tuple[str, int]]:
        """Findet nächsten verfügbaren Mitarbeiter aus Pool. Gibt (employee, new_position) zurück"""
        if not pool:
            return None

        pool_size = len(pool)
        for i in range(pool_size):
            candidate_idx = (start_pos + i) % pool_size
            candidate = pool[candidate_idx]
            if candidate not in absent and candidate not in excluded:
                return candidate, (candidate_idx + 1) % pool_size
        return None

    def _find_employee_with_catchup(self, pool: List[str], pool_key: str, start_pos: int,
                                     excluded: List[str], absent: List[str],
                                     catchup_queues: Dict[str, List[Dict]], current_tag_nr: int,
                                     pool_positions: Dict[str, int]) -> Tuple[Optional[str], int, List[str]]:
        """
        Sucht Mitarbeiter mit Berücksichtigung der Nachhol-Queue.
        Gibt (employee, new_position, skipped_employees) zurück.
        
        Logik:
        0. PRIORITÄTS-ESKALATION: Wer >= 4 Tage wartet, bekommt absolute Priorität
           - Wenn excluded wegen anderer Schicht am selben Tag: trotzdem einplanen!
        1. Prüfe ob der reguläre Kandidat (an start_pos) verfügbar ist
        2. Wenn ja → verwende diesen (normale Rotation)
        3. Wenn nein → prüfe Nachhol-Queue, dann andere aus Rotation
        4. Alle übersprungenen werden zur Queue hinzugefügt
        """
        skipped_employees: List[str] = []
        queue = catchup_queues.get(pool_key, [])
        
        if not pool:
            return None, start_pos, skipped_employees
        
        pool_size = len(pool)
        
        # 0. PRIORITÄTS-ESKALATION: Prüfe ob jemand in Queue zu lange wartet
        # Wichtig: Kopie der Queue durchlaufen, da wir sie modifizieren
        for i in range(len(queue) - 1, -1, -1):  # Rückwärts iterieren für sicheres pop()
            entry = queue[i]
            employee = entry["employee"]
            skipped_at = entry["skipped_at_tag"]
            waiting_days = current_tag_nr - skipped_at
            
            if waiting_days >= CATCHUP_PRIORITY_DAYS:
                # Dieser MA wartet zu lange - Priorität, aber excluded respektieren
                if employee not in absent and employee not in excluded:
                    # Nachhol-Priorität, aber kein MA soll VM+NM am selben Tag machen
                    queue.pop(i)
                    # Pool-Position NICHT ändern - reguläre Rotation bleibt erhalten
                    return employee, start_pos, skipped_employees
                else:
                    # Wirklich abwesend (Urlaub etc.) → Wartezeit zurücksetzen
                    entry["skipped_at_tag"] = current_tag_nr  # FIX: skipped_at zurücksetzen!
                    entry["available_from_tag"] = current_tag_nr + CATCHUP_DELAY_DAYS
        
        # 1. Prüfe zuerst den regulären Kandidaten (an start_pos)
        regular_candidate = pool[start_pos]
        if regular_candidate not in absent and regular_candidate not in excluded:
            # Regulärer Kandidat ist verfügbar! Normale Rotation.
            new_pos = (start_pos + 1) % pool_size
            return regular_candidate, new_pos, skipped_employees
        
        # 2. Regulärer Kandidat ist NICHT verfügbar → zur Nachhol-Queue hinzufügen
        already_in_queue = any(e["employee"] == regular_candidate for e in queue)
        if not already_in_queue:
            skipped_employees.append(regular_candidate)
        
        # 3. Jetzt prüfe Nachhol-Queue für Ersatz
        for i in range(len(queue) - 1, -1, -1):  # Rückwärts iterieren für sicheres pop()
            entry = queue[i]
            employee = entry["employee"]
            available_from = entry["available_from_tag"]
            
            if current_tag_nr >= available_from:
                # Dieser MA ist bereit für Nachhol-Versuch
                if employee not in absent and employee not in excluded:
                    # Verfügbar! Aus Queue entfernen und zurückgeben
                    queue.pop(i)
                    # Pool-Position NICHT ändern - reguläre Rotation bleibt erhalten
                    return employee, start_pos, skipped_employees
                # Nicht verfügbar, aber available_from erreicht → available_from aktualisieren
                elif employee in absent:
                    entry["available_from_tag"] = current_tag_nr + CATCHUP_DELAY_DAYS
                # Wenn excluded (nicht absent), warten wir auf Prioritäts-Eskalation
        
        # 4. Keine Nachhol-Kandidaten verfügbar → nächsten aus normaler Rotation suchen
        for i in range(1, pool_size):  # Start bei 1, da 0 (regulärer Kandidat) bereits geprüft
            candidate_idx = (start_pos + i) % pool_size
            candidate = pool[candidate_idx]
            
            if candidate not in absent and candidate not in excluded:
                # Gefunden! Gib zurück mit neuer Position
                new_pos = (candidate_idx + 1) % pool_size
                return candidate, new_pos, skipped_employees
            else:
                # Auch diesen als übersprungen markieren (wenn nicht schon in Queue)
                already_in_queue = any(e["employee"] == candidate for e in queue)
                if not already_in_queue and candidate not in skipped_employees:
                    skipped_employees.append(candidate)
        
        return None, start_pos, skipped_employees

    def create_planning(self) -> None:
        """Weiche: ruft je nach Toggle die klassische oder faire Planungslogik auf."""
        try:
            validation_result = self._validate_planning_inputs()
            if not validation_result:
                return

            # Von/Bis-Dropdowns der Abwesenheiten an den Planungszeitraum anpassen
            self._update_absence_date_options()

            if self.fair_mode_var.get():
                self._create_planning_fair(validation_result)
            else:
                self._create_planning_classic(validation_result)

            self.display_results()
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler bei der Planung: {e}")

    def _iterate_working_days(self, start_date: datetime):
        """Generator: liefert für jeden der 10 Arbeitstage Kontext-Informationen."""
        weekday_names_mofr = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag"]
        for working_day in range(WORKING_DAYS):
            weekday_in_week = working_day % 5
            week_number = working_day // 5
            current_date = start_date + timedelta(days=weekday_in_week + (week_number * 7))
            yield {
                "working_day": working_day,
                "current_date": current_date,
                "is_monday": weekday_in_week == 0,
                "is_wednesday": weekday_in_week == 2,
                "is_friday": weekday_in_week == 4,
                "weekday_german": weekday_names_mofr[weekday_in_week],
            }

    def _check_holiday(self, current_date: datetime) -> Optional[str]:
        """Prüft ob ein Datum ein Feiertag ist. Gibt den Notdienst-MA oder '' zurück, None wenn kein Feiertag."""
        current_date_str = current_date.strftime("%d.%m")
        for holiday in self.config.get("feiertage", []):
            if holiday["datum"] == current_date_str:
                return holiday.get("mitarbeiter", "")
        return None

    def _apply_holiday_or_first_day(self, holiday_employee: Optional[str], working_day: int,
                                     first_vm: str, first_nm: str, first_support: str):
        """Gibt (vm, nm, support) für Feiertage oder Tag 1 zurück, oder None wenn normaler Tag."""
        if holiday_employee is not None:
            if holiday_employee and holiday_employee != '-- Kein Notdienst --':
                return holiday_employee, holiday_employee, holiday_employee
            return "", "", ""
        if working_day == 0:
            return first_vm, first_nm, (first_support if first_support else "")
        return None

    def _create_planning_classic(self, validation_result) -> None:
        """Klassische Planung: Round-Robin mit Nachhol-Queue."""
        start_date, first_vm, first_nm, first_support = validation_result
        self.planning_result = []
        # Klassik-Modus liefert keine Fairness-Metrik
        self._fairness_baseline = None
        self._fairness_after = None
        self._fairness_base_counts = None
        pool_positions = self._initialize_pool_positions(first_vm, first_nm, first_support)
        catchup_queues: Dict[str, List[Dict]] = {
            "vm_alle": [], "nm_alle": [], "vm_support": []
        }

        for day in self._iterate_working_days(start_date):
            working_day = day["working_day"]
            absent_today = self._get_absent_for_day(
                working_day, day["current_date"], day["is_friday"],
                day["is_monday"], day["is_wednesday"], start_date
            )

            holiday_employee = self._check_holiday(day["current_date"])
            override = self._apply_holiday_or_first_day(
                holiday_employee, working_day, first_vm, first_nm, first_support
            )

            if override:
                vm_employee, nm_employee, support_employee = override
            else:
                vm_employee, nm_employee, support_employee = self._plan_day(
                    working_day, absent_today, pool_positions, catchup_queues
                )

            self.planning_result.append({
                "Datum": day["current_date"].strftime("%d.%m.%Y"),
                "Wochentag": day["weekday_german"],
                "Vormittag": vm_employee or "",
                "Nachmittag": nm_employee or "",
                "Support": support_employee or ""
            })

    def _create_planning_fair(self, validation_result) -> None:
        """Faire Planung: zählbasierte Auswahl mit Wochen-Quote und Optimierungspass.

        - Historie der letzten N Wochen als Startwert (gegen plan-internen Reset)
        - PRIMÄRES Kriterium ist die Einsatzzahl in der laufenden Woche
          ("niemand bekommt den 3. Einsatz, solange jemand Verfügbares 0 hat")
        - Tag 1 kann manuell vorgegeben werden (Felder ausgefüllt) oder wird fair mitgeplant
        - Nach der Erzeugung glättet ein Tausch-/Ersetzungspass die Verteilung weiter
        """
        start_date, first_vm, first_nm, first_support = validation_result
        self.planning_result = []

        weeks_back = self._get_fairness_horizon_weeks()
        base_counts: Dict[str, Dict[str, int]] = self._load_recent_history_counts(weeks_back)
        assignment_counts = {ma: dict(c) for ma, c in base_counts.items()}

        # Letzte Schicht pro MA (Tag-Nummer im aktuellen Plan); für Tiebreaker in _select_fairest
        last_shift_per_employee: Dict[str, int] = {}
        # Letzte Schicht pro MA UND Schichttyp (Tag-Nr); für 3-Tage-Cooldown auf gleiche Schicht
        last_shift_per_type: Dict[str, Dict[str, int]] = {}
        # Einsätze in der laufenden Woche (wird beim Wochenwechsel zurückgesetzt)
        week_counts: Dict[str, int] = {}
        # Fairness-Snapshot zum Vergleich (Zustand VOR dem aktuellen Plan)
        self._fairness_baseline = {ma: dict(c) for ma, c in base_counts.items()}
        self._fairness_base_counts = {ma: dict(c) for ma, c in base_counts.items()}

        # Tag 1 nur fixieren, wenn der Anwender etwas vorgegeben hat
        day0_locked = bool(first_vm or first_nm or first_support)

        def record(ma: str, key: str, tag_nr: int) -> None:
            assignment_counts.setdefault(ma, {"VM": 0, "NM": 0, "Support": 0})
            assignment_counts[ma][key] += 1
            last_shift_per_employee[ma] = tag_nr
            last_shift_per_type.setdefault(ma, {})[key] = tag_nr
            week_counts[ma] = week_counts.get(ma, 0) + 1

        for day in self._iterate_working_days(start_date):
            working_day = day["working_day"]
            if working_day == 5:
                week_counts = {}  # Neue Woche -> Wochen-Zähler zurücksetzen

            absent_today = self._get_absent_for_day(
                working_day, day["current_date"], day["is_friday"],
                day["is_monday"], day["is_wednesday"], start_date
            )

            holiday_employee = self._check_holiday(day["current_date"])
            if holiday_employee is not None:
                # Feiertag: ein MA übernimmt den ganzen Tag (oder Tag bleibt unbesetzt)
                if holiday_employee and holiday_employee != '-- Kein Notdienst --':
                    vm_employee = nm_employee = support_employee = holiday_employee
                    record(holiday_employee, "VM", working_day)
                    record(holiday_employee, "NM", working_day)
                else:
                    vm_employee = nm_employee = support_employee = ""
            else:
                presets = None
                if working_day == 0 and day0_locked:
                    presets = {"VM": first_vm, "NM": first_nm, "Support": first_support}
                vm_employee, nm_employee, support_employee = self._plan_day_fair(
                    working_day, absent_today, assignment_counts,
                    last_shift_per_employee, last_shift_per_type, week_counts,
                    record, presets=presets
                )

            self.planning_result.append({
                "Datum": day["current_date"].strftime("%d.%m.%Y"),
                "Wochentag": day["weekday_german"],
                "Vormittag": vm_employee or "",
                "Nachmittag": nm_employee or "",
                "Support": support_employee or ""
            })

        # Optimierungspass: gültige Tausch-/Ersetzungsoperationen, die die Verteilung verbessern
        self._improve_plan_fairness(start_date, base_counts, day0_locked)

        # Endstand für die Fairness-Metrik aus dem finalen Plan neu berechnen
        self._fairness_after = self._combine_counts(
            base_counts, self._count_plan_assignments(self.planning_result)
        )

    def _load_recent_history_counts(self, weeks_back: int) -> Dict[str, Dict[str, int]]:
        """Initialisiert Schicht-Zähler aus der gespeicherten Historie für den Fair-Modus.

        weeks_back: Anzahl Wochen rückwärts ab heute, die berücksichtigt werden.
                    0 = keine Historie (leerer Startwert), -1 = komplette Historie.
        """
        counts: Dict[str, Dict[str, int]] = {}
        if weeks_back == 0:
            return counts

        cutoff: Optional[datetime] = None
        if weeks_back > 0:
            cutoff = datetime.now() - timedelta(weeks=weeks_back)

        for plan in self.saved_plans:
            for entry in plan.get("entries", []):
                try:
                    entry_date = datetime.strptime(entry["Datum"], "%d.%m.%Y")
                except (ValueError, KeyError):
                    continue
                if cutoff and entry_date < cutoff:
                    continue
                # Support == Vormittag ist nur Anzeige-Duplizierung, kein echter Einsatz
                vm_ma = entry.get("Vormittag", "").strip()
                for shift_type, key in [("Vormittag", "VM"), ("Nachmittag", "NM"), ("Support", "Support")]:
                    ma = entry.get(shift_type, "").strip()
                    if not ma:
                        continue
                    if key == "Support" and ma == vm_ma:
                        continue
                    counts.setdefault(ma, {"VM": 0, "NM": 0, "Support": 0})
                    counts[ma][key] += 1
        return counts

    @staticmethod
    def _compute_total_load(ma: str, counts: Dict[str, Dict[str, int]]) -> int:
        """Summe aller Schichten (VM+NM+Support) eines Mitarbeiters."""
        c = counts.get(ma)
        if not c:
            return 0
        return c.get("VM", 0) + c.get("NM", 0) + c.get("Support", 0)

    def _fair_weight(self, ma: str) -> float:
        """Gewichtungsfaktor für die Soll-Verteilung (Pool B zählt 1,5-fach)."""
        if ma in self.config.get("pool_vm_teilweise", []):
            return FAIR_POOL_B_WEIGHT
        return 1.0

    @staticmethod
    def _compute_fairness_metrics(counts: Dict[str, Dict[str, int]]) -> Dict[str, object]:
        """Berechnet Fairness-Kennzahlen (σ, min, max, spread) aus der Gesamtlast pro MA.

        Gibt {} zurück, wenn keine Daten vorliegen.
        """
        if not counts:
            return {}
        loads = [(ma, c.get("VM", 0) + c.get("NM", 0) + c.get("Support", 0)) for ma, c in counts.items()]
        loads = [(ma, total) for ma, total in loads if total > 0]
        if not loads:
            return {}
        values = [t for _, t in loads]
        n = len(values)
        mean = sum(values) / n
        variance = sum((v - mean) ** 2 for v in values) / n
        sigma = variance ** 0.5
        min_ma, min_v = min(loads, key=lambda x: x[1])
        max_ma, max_v = max(loads, key=lambda x: x[1])
        return {
            "sigma": sigma,
            "min_ma": min_ma,
            "min_value": min_v,
            "max_ma": max_ma,
            "max_value": max_v,
            "spread": max_v - min_v,
            "mean": mean,
            "n": n,
        }

    def _select_fairest(self, pool: List[str], shift_key: str,
                        excluded: List[str], absent: List[str],
                        assignment_counts: Dict[str, Dict[str, int]],
                        yesterday_assigned: Optional[str],
                        last_shift_per_employee: Optional[Dict[str, int]] = None,
                        current_tag_nr: int = 0,
                        yesterday_anywhere: Optional[List[str]] = None,
                        last_shift_per_type: Optional[Dict[str, Dict[str, int]]] = None,
                        week_counts: Optional[Dict[str, int]] = None) -> Optional[str]:
        """Wählt den verfügbaren MA mit den wenigsten Einsätzen (Fair-Modus).

        Mehrstufiger Sort-Key (kleiner = besser):
        1. week_count   — Einsätze in der LAUFENDEN Woche (Wochen-Quote: niemand bekommt
                          den 3. Einsatz, solange ein Verfügbarer noch 0/1 hat)
        2. count        — Schichten dieses Typs (primäre Langzeit-Fairness)
        3. total_load   — Summe VM+NM+Support (gleicht Multi-Pool-MA aus)
        4. yesterday_penalty — 1 wenn gestern irgendwo eingeplant, sonst 0 (verteilt Last über Woche)
        5. -days_since_same_type — länger nicht denselben Typ = besser (verteilt Schichttyp gleichmäßig)
        6. -days_since_any  — länger gar nicht dran = besser (Last-Resort gegen Pool-Index-Bias)
        7. idx          — Pool-Index nur als deterministischer Last-Resort

        Drei-Stufen-Cooldown beim selben Schichttyp:
          1. Strikt:  letzter Einsatz desselben Typs >= FAIR_MIN_GAP_SAME_SHIFT Tage her
          2. Locker:  nicht direkt gestern derselbe Typ (yesterday_assigned)
          3. Fallback: alles erlaubt (wenn Pool ausgeschöpft)
        """
        last_shift_map = last_shift_per_employee or {}
        last_type_map = last_shift_per_type or {}
        yesterday_set = set(yesterday_anywhere or [])
        week_map = week_counts or {}
        pool_f_set = set(self.config.get("pool_mo_mi_abwesend", []))

        def build_candidates(min_gap_same_type: int, allow_consecutive: bool) -> List[Tuple[int, int, int, int, int, int, int, str]]:
            result = []
            for idx, member in enumerate(pool):
                if member in absent or member in excluded:
                    continue
                # Pool F (Mo/Mi-frei): harte Grenze von max. 1 Einsatz pro Woche
                if member in pool_f_set and week_map.get(member, 0) >= 1:
                    continue
                if not allow_consecutive and member == yesterday_assigned:
                    continue
                # Cooldown: letzter Einsatz desselben Schichttyps muss >= min_gap_same_type Arbeitstage her sein
                last_same_type = last_type_map.get(member, {}).get(shift_key, -10_000)
                if min_gap_same_type > 0:
                    if current_tag_nr - last_same_type < min_gap_same_type:
                        continue
                # Gewichtete Zähler: Pool-B-Einsätze zählen 1,5-fach (Soll-Verteilung)
                weight = self._fair_weight(member)
                week_count = week_map.get(member, 0) * weight
                count = assignment_counts.get(member, {}).get(shift_key, 0) * weight
                total_load = self._compute_total_load(member, assignment_counts) * weight
                yesterday_penalty = 1 if member in yesterday_set else 0
                neg_days_since_same = -(current_tag_nr - last_same_type)
                last_any = last_shift_map.get(member, -10_000)
                neg_days_since_any = -(current_tag_nr - last_any)
                result.append((week_count, count, total_load, yesterday_penalty,
                               neg_days_since_same, neg_days_since_any, idx, member))
            return result

        # Stufe 1: strikter Cooldown (3 Arbeitstage Pause beim selben Schichttyp) + kein gleicher Dienst gestern
        candidates = build_candidates(min_gap_same_type=FAIR_MIN_GAP_SAME_SHIFT, allow_consecutive=False)
        # Stufe 2: nur "nicht gestern derselbe Typ"
        if not candidates:
            candidates = build_candidates(min_gap_same_type=0, allow_consecutive=False)
        # Stufe 3: alles erlaubt
        if not candidates:
            candidates = build_candidates(min_gap_same_type=0, allow_consecutive=True)
        if not candidates:
            return None

        candidates.sort()
        return candidates[0][-1]

    def _pool_f_week_capped(self, tag_nr: int) -> set:
        """Pool-F-MA, die in der laufenden Woche bereits einen Einsatz im aktuellen Plan haben.

        Pool F = Mo/Mi nicht verfügbar; für diese MA gilt zusätzlich: max. 1 Einsatz pro Woche.
        """
        pool_f = set(self.config.get("pool_mo_mi_abwesend", []))
        if not pool_f:
            return set()
        week_start = (tag_nr // 5) * 5
        capped = set()
        for i in range(week_start, min(tag_nr, len(self.planning_result))):
            e = self.planning_result[i]
            vm = e.get("Vormittag", "").strip()
            for shift_type in ("Vormittag", "Nachmittag", "Support"):
                ma = e.get(shift_type, "").strip()
                if ma in pool_f and not (shift_type == "Support" and ma == vm):
                    capped.add(ma)
        return capped

    def _plan_day(self, tag_nr: int, absent_today: Dict[str, List[str]],
                  pool_positions: Dict[str, int],
                  catchup_queues: Dict[str, List[Dict]]) -> Tuple[Optional[str], Optional[str], str]:
        """Plant einen einzelnen Tag mit Nachhol-Queue. Gibt (vm_employee, nm_employee, support_employee) zurück"""
        yesterday_nm = self.planning_result[tag_nr - 1]["Nachmittag"] if tag_nr > 0 else None
        forbidden_vm = [yesterday_nm] if yesterday_nm else []
        used_today = []

        # Pool F (Mo/Mi-frei): max. 1 Einsatz pro Woche — bereits Eingeteilte sperren
        week_capped = self._pool_f_week_capped(tag_nr)
        if week_capped:
            absent_today = {
                "VM": sorted(set(absent_today["VM"]) | week_capped),
                "NM": sorted(set(absent_today["NM"]) | week_capped),
            }

        # Vormittag planen (mit Nachhol-Queue)
        vm_employee, new_vm_pos, vm_skipped = self._find_employee_with_catchup(
            self.config["pool_vm_alle"],
            "vm_alle",
            pool_positions["vm_alle"],
            forbidden_vm,
            absent_today["VM"],
            catchup_queues,
            tag_nr,
            pool_positions
        )
        
        pool_positions["vm_alle"] = new_vm_pos
        support_employee = ""

        # Übersprungene zur Queue hinzufügen
        for skipped in vm_skipped:
            catchup_queues["vm_alle"].append({
                "employee": skipped,
                "skipped_at_tag": tag_nr,
                "available_from_tag": tag_nr + CATCHUP_DELAY_DAYS
            })

        if vm_employee:
            used_today.append(vm_employee)

            # Support nötig?
            # 1. Immer wenn VM aus Pool B (braucht Unterstützung)
            # 2. Montag in Woche 2 (working_day == 5) - höheres Aufkommen (wenn Toggle nicht aktiv)
            monday_needs_support = (tag_nr == 5) and not self.monday_normal_var.get()
            needs_support = (vm_employee in self.config["pool_vm_teilweise"]) or monday_needs_support
            
            if needs_support:
                support_employee, new_support_pos, support_skipped = self._find_employee_with_catchup(
                    self.config["pool_vm_support"],
                    "vm_support",
                    pool_positions["vm_support"],
                    [vm_employee] + forbidden_vm,
                    absent_today["VM"],
                    catchup_queues,
                    tag_nr,
                    pool_positions
                )
                pool_positions["vm_support"] = new_support_pos
                
                # Übersprungene Support-MA zur Queue
                for skipped in support_skipped:
                    catchup_queues["vm_support"].append({
                        "employee": skipped,
                        "skipped_at_tag": tag_nr,
                        "available_from_tag": tag_nr + CATCHUP_DELAY_DAYS
                    })
                
                if support_employee:
                    used_today.append(support_employee)
            else:
                # Kein separater Support nötig → VM-MA auch in Support-Spalte eintragen (Übersichtlichkeit)
                support_employee = vm_employee

        # Nachmittag planen (mit Nachhol-Queue)
        nm_employee, new_nm_pos, nm_skipped = self._find_employee_with_catchup(
            self.config["pool_nm_alle"],
            "nm_alle",
            pool_positions["nm_alle"],
            used_today,
            absent_today["NM"],
            catchup_queues,
            tag_nr,
            pool_positions
        )
        
        pool_positions["nm_alle"] = new_nm_pos

        # Übersprungene zur Queue hinzufügen
        for skipped in nm_skipped:
            catchup_queues["nm_alle"].append({
                "employee": skipped,
                "skipped_at_tag": tag_nr,
                "available_from_tag": tag_nr + CATCHUP_DELAY_DAYS
            })

        return vm_employee, nm_employee, support_employee

    def _plan_day_fair(self, tag_nr: int, absent_today: Dict[str, List[str]],
                       assignment_counts: Dict[str, Dict[str, int]],
                       last_shift_per_employee: Dict[str, int],
                       last_shift_per_type: Dict[str, Dict[str, int]],
                       week_counts: Dict[str, int],
                       record,
                       presets: Optional[Dict[str, str]] = None) -> Tuple[Optional[str], Optional[str], str]:
        """Plant einen einzelnen Tag mit zählbasierter Fairness-Logik.

        presets: optionale manuelle Vorgaben für Tag 1 ({"VM":..., "NM":..., "Support":...}).
                 Leere Vorgaben werden fair ergänzt (z.B. nur VM vorgegeben -> NM fair geplant).
        """
        presets = presets or {}
        yesterday_vm = self.planning_result[tag_nr - 1]["Vormittag"] if tag_nr > 0 else None
        yesterday_nm = self.planning_result[tag_nr - 1]["Nachmittag"] if tag_nr > 0 else None
        yesterday_support = self.planning_result[tag_nr - 1]["Support"] if tag_nr > 0 else None
        forbidden_vm = [yesterday_nm] if yesterday_nm else []
        used_today: List[str] = []

        # "Gestern irgendwo eingeteilt" -> Soft-Penalty in _select_fairest
        yesterday_anywhere: List[str] = []
        if tag_nr > 0:
            yesterday_anywhere = [
                v for v in (yesterday_vm, yesterday_nm, yesterday_support) if v
            ]

        # Vormittag
        if presets.get("VM"):
            vm_employee = presets["VM"]
        else:
            vm_employee = self._select_fairest(
                self.config["pool_vm_alle"], "VM",
                forbidden_vm, absent_today["VM"],
                assignment_counts, yesterday_vm,
                last_shift_per_employee=last_shift_per_employee,
                current_tag_nr=tag_nr,
                yesterday_anywhere=yesterday_anywhere,
                last_shift_per_type=last_shift_per_type,
                week_counts=week_counts,
            )

        support_employee = ""
        if vm_employee:
            used_today.append(vm_employee)
            record(vm_employee, "VM", tag_nr)

            monday_needs_support = (tag_nr == 5) and not self.monday_normal_var.get()
            needs_support = (vm_employee in self.config["pool_vm_teilweise"]) or monday_needs_support

            preset_support = presets.get("Support", "")
            if preset_support and preset_support != vm_employee:
                support_employee = preset_support
                used_today.append(support_employee)
                record(support_employee, "Support", tag_nr)
            elif needs_support:
                support_employee = self._select_fairest(
                    self.config["pool_vm_support"], "Support",
                    [vm_employee] + forbidden_vm, absent_today["VM"],
                    assignment_counts, yesterday_support,
                    last_shift_per_employee=last_shift_per_employee,
                    current_tag_nr=tag_nr,
                    yesterday_anywhere=yesterday_anywhere,
                    last_shift_per_type=last_shift_per_type,
                    week_counts=week_counts,
                )
                if support_employee:
                    used_today.append(support_employee)
                    record(support_employee, "Support", tag_nr)
            else:
                support_employee = vm_employee

        # Nachmittag
        if presets.get("NM"):
            nm_employee = presets["NM"]
        else:
            nm_employee = self._select_fairest(
                self.config["pool_nm_alle"], "NM",
                used_today, absent_today["NM"],
                assignment_counts, yesterday_nm,
                last_shift_per_employee=last_shift_per_employee,
                current_tag_nr=tag_nr,
                yesterday_anywhere=yesterday_anywhere,
                last_shift_per_type=last_shift_per_type,
                week_counts=week_counts,
            )
        if nm_employee:
            record(nm_employee, "NM", tag_nr)

        return vm_employee, nm_employee, support_employee

    def _get_fairness_horizon_weeks(self) -> int:
        """Liefert die aktuell ausgewählten Wochen für den Fairness-Horizont.

        Liest die GUI-Var, fällt auf die persistierte Einstellung oder den Default zurück.
        """
        label = None
        if hasattr(self, 'fairness_horizon_var'):
            try:
                label = self.fairness_horizon_var.get()
            except Exception:
                label = None
        if not label:
            label = getattr(self, '_fairness_horizon_label', FAIRNESS_HORIZON_DEFAULT_LABEL)
        for opt_label, opt_weeks in FAIRNESS_HORIZON_OPTIONS:
            if opt_label == label:
                return opt_weeks
        # Unbekanntes Label -> Default
        for opt_label, opt_weeks in FAIRNESS_HORIZON_OPTIONS:
            if opt_label == FAIRNESS_HORIZON_DEFAULT_LABEL:
                return opt_weeks
        return 13

    # -------------------- Fairness-Helfer & Optimierungspass --------------------

    @staticmethod
    def _count_plan_assignments(entries: List[Dict[str, str]]) -> Dict[str, Dict[str, int]]:
        """Zählt Einsätze im Plan (Support == VM zählt nicht doppelt)."""
        counts: Dict[str, Dict[str, int]] = {}
        for e in entries:
            vm = e.get("Vormittag", "").strip()
            for shift_type, key in [("Vormittag", "VM"), ("Nachmittag", "NM"), ("Support", "Support")]:
                ma = e.get(shift_type, "").strip()
                if not ma:
                    continue
                if key == "Support" and ma == vm:
                    continue
                counts.setdefault(ma, {"VM": 0, "NM": 0, "Support": 0})
                counts[ma][key] += 1
        return counts

    @staticmethod
    def _combine_counts(base: Dict[str, Dict[str, int]],
                        plan: Dict[str, Dict[str, int]]) -> Dict[str, Dict[str, int]]:
        """Addiert zwei Zähl-Dictionaries (Historie + aktueller Plan)."""
        out = {ma: dict(c) for ma, c in base.items()}
        for ma, c in plan.items():
            out.setdefault(ma, {"VM": 0, "NM": 0, "Support": 0})
            for key, val in c.items():
                out[ma][key] = out[ma].get(key, 0) + val
        return out

    def _collect_day_absences(self, start_date: datetime) -> List[Dict[str, List[str]]]:
        """Berechnet für alle 10 Arbeitstage die Abwesenheitslisten (VM/NM)."""
        absents = []
        for day in self._iterate_working_days(start_date):
            absents.append(self._get_absent_for_day(
                day["working_day"], day["current_date"], day["is_friday"],
                day["is_monday"], day["is_wednesday"], start_date
            ))
        return absents

    def _plan_objective(self, base_counts: Dict[str, Dict[str, int]],
                        absents: List[Dict[str, List[str]]]) -> Tuple[int, int, float, int]:
        """Bewertet den aktuellen Plan: (Wochen-Spread-Summe, Gesamt-Spread, Sigma, Folgetag-Strafe).

        Kleiner = besser. Wochen-Spread zählt nur Mitarbeiter, die in der Woche
        überhaupt an mindestens einem Tag verfügbar sind. Die Folgetag-Strafe zählt,
        wie oft derselbe MA an zwei direkt aufeinanderfolgenden Tagen eingeteilt ist.
        """
        entries = self.planning_result
        employees = self._get_all_employees()

        week_spread_sum = 0
        for w in (0, 1):
            idxs = [i for i in range(w * 5, min((w + 1) * 5, len(entries)))]
            if not idxs:
                continue
            available = []
            for ma in employees:
                for i in idxs:
                    a = absents[i] if i < len(absents) else {"VM": [], "NM": []}
                    if ma not in a["VM"] or ma not in a["NM"]:
                        available.append(ma)
                        break
            loads = {ma: 0 for ma in available}
            for i in idxs:
                e = entries[i]
                vm = e.get("Vormittag", "").strip()
                for shift_type in ("Vormittag", "Nachmittag", "Support"):
                    ma = e.get(shift_type, "").strip()
                    if not ma:
                        continue
                    if shift_type == "Support" and ma == vm:
                        continue
                    if ma in loads:
                        loads[ma] += 1
            if loads:
                # Gewichtete Wochenlast (Pool B zählt 1,5-fach)
                values = [loads[ma] * self._fair_weight(ma) for ma in loads]
                week_spread_sum += round(max(values) - min(values), 3)

        # Gewichtete Gesamtlast (Historie + Plan) über alle Pool-Mitarbeiter
        totals = []
        plan_counts = self._count_plan_assignments(entries)
        combined = self._combine_counts(base_counts, plan_counts)
        for ma in employees:
            totals.append(self._compute_total_load(ma, combined) * self._fair_weight(ma))
        if totals:
            spread_total = round(max(totals) - min(totals), 3)
            mean = sum(totals) / len(totals)
            sigma = (sum((v - mean) ** 2 for v in totals) / len(totals)) ** 0.5
        else:
            spread_total, sigma = 0, 0.0

        # Folgetag-Strafe: derselbe MA an zwei aufeinanderfolgenden Tagen
        repeat_penalty = 0
        prev_assigned: set = set()
        for e in entries:
            vm = e.get("Vormittag", "").strip()
            today_assigned = set()
            for shift_type in ("Vormittag", "Nachmittag", "Support"):
                ma = e.get(shift_type, "").strip()
                if ma and not (shift_type == "Support" and ma == vm):
                    today_assigned.add(ma)
            repeat_penalty += len(today_assigned & prev_assigned)
            prev_assigned = today_assigned

        return (week_spread_sum, spread_total, round(sigma, 4), repeat_penalty)

    def _count_soft_issues(self) -> int:
        """Zählt 'weiche' Verstöße im Plan (für das Gate des Optimierungspasses):

        - dieselbe Person an zwei direkt aufeinanderfolgenden Arbeitstagen
        - derselbe Schichttyp derselben Person innerhalb von FAIR_MIN_GAP_SAME_SHIFT Arbeitstagen
        """
        entries = self.planning_result
        issues = 0
        prev_assigned: set = set()
        last_type: Dict[str, Dict[str, int]] = {}
        for i, e in enumerate(entries):
            vm = e.get("Vormittag", "").strip()
            today: set = set()
            for shift_type, key in (("Vormittag", "VM"), ("Nachmittag", "NM"), ("Support", "Support")):
                ma = e.get(shift_type, "").strip()
                if not ma or (shift_type == "Support" and ma == vm):
                    continue
                today.add(ma)
                lt = last_type.get(ma, {}).get(key)
                if lt is not None and i - lt < FAIR_MIN_GAP_SAME_SHIFT:
                    issues += 1
                last_type.setdefault(ma, {})[key] = i
            issues += len(today & prev_assigned)
            prev_assigned = today
        return issues

    def _improve_plan_fairness(self, start_date: datetime,
                               base_counts: Dict[str, Dict[str, int]],
                               day0_locked: bool) -> None:
        """Optimierungspass: probiert gültige Ersetzungen und Tauschoperationen aus
        und übernimmt nur Änderungen, die die Fairness messbar verbessern.

        Gültigkeit wird doppelt abgesichert:
        - Plan-Prüfung: keine neuen harten Fehler
        - Soft-Gate: keine neuen Folgetags-Einsätze oder kurz aufeinanderfolgende
          gleiche Schichttypen (sonst 'optimiert' der Pass Häufungen herbei)
        """
        entries = self.planning_result
        if not entries:
            return

        absents = self._collect_day_absences(start_date)

        locked_rows = set()
        for i, day in enumerate(self._iterate_working_days(start_date)):
            if self._check_holiday(day["current_date"]) is not None:
                locked_rows.add(i)
        if day0_locked:
            locked_rows.add(0)

        role_pools = {
            "Vormittag": self.config.get("pool_vm_alle", []),
            "Nachmittag": self.config.get("pool_nm_alle", []),
            "Support": self.config.get("pool_vm_support", []),
        }
        teilweise = self.config.get("pool_vm_teilweise", [])

        def fehler_count() -> int:
            return sum(1 for v in self._validate_current_plan() if v["schwere"] == "fehler")

        allowed_fehler = fehler_count()
        allowed_soft = self._count_soft_issues()
        best_obj = self._plan_objective(base_counts, absents)

        for _ in range(FAIR_IMPROVE_MAX_ROUNDS):
            improved = False

            # 1) Ersetzungen: einzelne Zuweisung durch anderen Pool-MA ersetzen
            for i, e in enumerate(entries):
                if i in locked_rows:
                    continue
                for role, pool in role_pools.items():
                    cur = e.get(role, "").strip()
                    if not cur:
                        continue
                    # Anzeige-Duplikat (Support == VM) ist kein echter Support -> gesperrt
                    if role == "Support" and cur == e.get("Vormittag", "").strip():
                        continue
                    sup_is_dup = (role == "Vormittag" and e.get("Support", "").strip() == cur)
                    old_sup = e.get("Support", "")

                    best_candidate = None
                    for q in pool:
                        if q == cur:
                            continue
                        # Bei VM-Ersetzung mit Duplikat-Support: Pool-B-Ersatz bräuchte echten Support
                        if sup_is_dup and q in teilweise:
                            continue
                        e[role] = q
                        if sup_is_dup:
                            e["Support"] = q
                        if (fehler_count() <= allowed_fehler
                                and self._count_soft_issues() <= allowed_soft):
                            obj = self._plan_objective(base_counts, absents)
                            if obj < best_obj:
                                best_obj = obj
                                best_candidate = q
                        e[role] = cur
                        if sup_is_dup:
                            e["Support"] = old_sup

                    if best_candidate:
                        e[role] = best_candidate
                        if sup_is_dup:
                            e["Support"] = best_candidate
                        allowed_soft = min(allowed_soft, self._count_soft_issues())
                        improved = True

            # 2) Tausch derselben Rolle zwischen zwei Tagen (balanciert Wochen aus)
            for role in role_pools:
                for i in range(len(entries)):
                    if i in locked_rows:
                        continue
                    for j in range(i + 1, len(entries)):
                        if j in locked_rows:
                            continue
                        a = entries[i].get(role, "").strip()
                        b = entries[j].get(role, "").strip()
                        if not a or not b or a == b:
                            continue
                        if role == "Support" and (
                            a == entries[i].get("Vormittag", "").strip()
                            or b == entries[j].get("Vormittag", "").strip()
                        ):
                            continue
                        dup_i = (role == "Vormittag" and entries[i].get("Support", "").strip() == a)
                        dup_j = (role == "Vormittag" and entries[j].get("Support", "").strip() == b)

                        entries[i][role], entries[j][role] = b, a
                        if dup_i:
                            entries[i]["Support"] = b
                        if dup_j:
                            entries[j]["Support"] = a

                        accepted = False
                        if (fehler_count() <= allowed_fehler
                                and self._count_soft_issues() <= allowed_soft):
                            obj = self._plan_objective(base_counts, absents)
                            if obj < best_obj:
                                best_obj = obj
                                accepted = True
                                improved = True
                                allowed_soft = min(allowed_soft, self._count_soft_issues())

                        if not accepted:
                            entries[i][role], entries[j][role] = a, b
                            if dup_i:
                                entries[i]["Support"] = a
                            if dup_j:
                                entries[j]["Support"] = b

            if not improved:
                break

    # -------------------- Plan-Prüfung (automatische Regelkontrolle) --------------------

    def _validate_current_plan(self) -> List[Dict]:
        """Prüft den aktuellen Plan gegen alle Regeln.

        Rückgabe: Liste von {"row": int|None, "schwere": "fehler"|"hinweis", "text": str}
        """
        violations: List[Dict] = []
        entries = self.planning_result
        if not entries:
            return violations

        known = set(self._get_all_employees())
        try:
            start_date = datetime.strptime(entries[0].get("Datum", ""), "%d.%m.%Y")
        except ValueError:
            return violations

        monday_normal = False
        if hasattr(self, 'monday_normal_var'):
            try:
                monday_normal = bool(self.monday_normal_var.get())
            except Exception:
                monday_normal = False
        teilweise = self.config.get("pool_vm_teilweise", [])

        def add(row: Optional[int], schwere: str, text: str) -> None:
            violations.append({"row": row, "schwere": schwere, "text": text})

        absents_per_row: List[Dict[str, List[str]]] = []
        prev_nm: Optional[str] = None

        for i, e in enumerate(entries):
            datum_str = e.get("Datum", "")
            try:
                d = datetime.strptime(datum_str, "%d.%m.%Y")
            except ValueError:
                absents_per_row.append({"VM": [], "NM": []})
                continue

            vm = e.get("Vormittag", "").strip()
            nm = e.get("Nachmittag", "").strip()
            sup = e.get("Support", "").strip()
            wd = d.weekday()
            absent = self._get_absent_for_day(i, d, wd == 4, wd == 0, wd == 2, start_date)
            absents_per_row.append(absent)

            # Unbekannte Kürzel (Tippfehler-Schutz) — gilt auch an Feiertagen
            for role, val in (("VM", vm), ("NM", nm), ("Support", sup)):
                if val and val not in known:
                    add(i, "hinweis", f"{datum_str}: Unbekanntes Kürzel '{val}' ({role})")

            holiday = self._check_holiday(d)
            if holiday is not None:
                # Feiertag: ein MA ganztags (oder bewusst unbesetzt) — übrige Regeln greifen nicht
                prev_nm = nm or None
                continue

            if not vm:
                add(i, "fehler", f"{datum_str}: Vormittag unbesetzt")
            if not nm:
                add(i, "fehler", f"{datum_str}: Nachmittag unbesetzt")
            if vm and vm == nm:
                add(i, "fehler", f"{datum_str}: {vm} ist VM und NM am selben Tag")
            if sup and sup == nm and sup != vm:
                add(i, "fehler", f"{datum_str}: {sup} ist Support und NM am selben Tag")
            if vm and prev_nm and vm == prev_nm:
                add(i, "fehler", f"{datum_str}: {vm} hatte am Vortag Nachmittag (Ruhezeit verletzt)")
            prev_work_date = d - timedelta(days=3 if d.weekday() == 0 else 1)
            prev_full_absent = self._full_day_absent_on(prev_work_date)
            if vm and vm in prev_full_absent:
                add(i, "fehler",
                    f"{datum_str}: {vm} war am Vortag ganztags abwesend — "
                    "kein Vormittag direkt nach Abwesenheit")
            elif vm and vm in absent["VM"]:
                add(i, "fehler", f"{datum_str}: {vm} ist vormittags abwesend/ausgeschlossen")
            if sup and sup != vm and sup in prev_full_absent:
                add(i, "fehler",
                    f"{datum_str}: Support {sup} war am Vortag ganztags abwesend — "
                    "kein Vormittagseinsatz direkt nach Abwesenheit")
            elif sup and sup != vm and sup in absent["VM"]:
                add(i, "fehler", f"{datum_str}: Support {sup} ist vormittags abwesend/ausgeschlossen")
            next_work_date = d + timedelta(days=3 if d.weekday() == 4 else 1)
            next_day_vm_absent, _ = self._absence_sets_for_date(next_work_date)
            if nm and nm in next_day_vm_absent:
                add(i, "fehler",
                    f"{datum_str}: {nm} ist am Folgetag ({next_work_date.strftime('%d.%m.%Y')}) "
                    "abwesend — kein Nachmittag am Vortag")
            elif nm and nm in absent["NM"]:
                add(i, "fehler", f"{datum_str}: {nm} ist nachmittags abwesend/ausgeschlossen")
            if vm and vm in teilweise and (not sup or sup == vm):
                add(i, "fehler", f"{datum_str}: {vm} (Pool B) ohne Support eingeteilt")
            if i == 5 and not monday_normal and vm and (not sup or sup == vm):
                add(i, "hinweis", f"{datum_str}: Montag Woche 2 ohne separaten Support")

            prev_nm = nm or None

        # Hinweis: gleicher Schichttyp in kurzem Abstand (Soll-Abstand >= FAIR_MIN_GAP_SAME_SHIFT)
        last_type_seen: Dict[str, Dict[str, int]] = {}
        type_labels = {"VM": "Vormittag", "NM": "Nachmittag", "Support": "Support"}
        for i, e in enumerate(entries):
            datum_str = e.get("Datum", "")
            try:
                d = datetime.strptime(datum_str, "%d.%m.%Y")
            except ValueError:
                continue
            if self._check_holiday(d) is not None:
                continue
            vm = e.get("Vormittag", "").strip()
            for shift_type, key in (("Vormittag", "VM"), ("Nachmittag", "NM"), ("Support", "Support")):
                ma = e.get(shift_type, "").strip()
                if not ma or (shift_type == "Support" and ma == vm):
                    continue
                prev_i = last_type_seen.get(ma, {}).get(key)
                if prev_i is not None and i - prev_i < FAIR_MIN_GAP_SAME_SHIFT:
                    gap = i - prev_i
                    add(i, "hinweis",
                        f"{datum_str}: {ma} hat erneut {type_labels[key]} nach nur {gap} Arbeitstag(en) "
                        f"(Soll-Abstand: {FAIR_MIN_GAP_SAME_SHIFT})")
                last_type_seen.setdefault(ma, {})[key] = i

        # Pool-F-Limit: max. 1 Einsatz pro Woche
        pool_f = set(self.config.get("pool_mo_mi_abwesend", []))
        if pool_f:
            for w in (0, 1):
                idxs = [i for i in range(w * 5, min((w + 1) * 5, len(entries)))]
                counts_f: Dict[str, int] = {}
                for i in idxs:
                    e = entries[i]
                    vm = e.get("Vormittag", "").strip()
                    for shift_type in ("Vormittag", "Nachmittag", "Support"):
                        ma = e.get(shift_type, "").strip()
                        if not ma or ma not in pool_f or (shift_type == "Support" and ma == vm):
                            continue
                        counts_f[ma] = counts_f.get(ma, 0) + 1
                        if counts_f[ma] == 2:
                            add(i, "fehler",
                                f"{e.get('Datum', '')}: {ma} (Pool F) hat mehr als 1 Einsatz "
                                f"in Woche {w + 1} — erlaubt ist max. 1 pro Woche")

        # Wochenbalance-Hinweis: große Spreizung innerhalb einer Woche
        employees = self._get_all_employees()
        for w in (0, 1):
            idxs = [i for i in range(w * 5, min((w + 1) * 5, len(entries)))]
            if not idxs:
                continue
            available = []
            for ma in employees:
                for i in idxs:
                    a = absents_per_row[i] if i < len(absents_per_row) else {"VM": [], "NM": []}
                    if ma not in a["VM"] or ma not in a["NM"]:
                        available.append(ma)
                        break
            loads = {ma: 0 for ma in available}
            for i in idxs:
                e = entries[i]
                vm = e.get("Vormittag", "").strip()
                for shift_type in ("Vormittag", "Nachmittag", "Support"):
                    ma = e.get(shift_type, "").strip()
                    if not ma or (shift_type == "Support" and ma == vm):
                        continue
                    if ma in loads:
                        loads[ma] += 1
            if loads:
                values = list(loads.values())
                spread = max(values) - min(values)
                if spread >= WEEK_BALANCE_HINT_SPREAD:
                    min_ma = min(loads, key=lambda m: loads[m])
                    max_ma = max(loads, key=lambda m: loads[m])
                    add(None, "hinweis",
                        f"Woche {w + 1}: ungleiche Verteilung — {max_ma} hat {loads[max_ma]}, "
                        f"{min_ma} nur {loads[min_ma]} Einsätze")

        return violations

    def _run_plan_checks(self) -> None:
        """Führt die Plan-Prüfung aus und aktualisiert Prüf-Panel, Zeilenfarben,
        Wochenlast-Panel und Fairness-Metrik."""
        if not hasattr(self, 'check_tree'):
            return

        violations = self._validate_current_plan() if self.planning_result else []
        self._plan_violations = violations

        # Prüf-Panel füllen
        for item in self.check_tree.get_children():
            self.check_tree.delete(item)
        fehler_n = sum(1 for v in violations if v["schwere"] == "fehler")
        hinweis_n = sum(1 for v in violations if v["schwere"] == "hinweis")
        for v in violations:
            label = "Fehler" if v["schwere"] == "fehler" else "Hinweis"
            self.check_tree.insert("", tk.END, values=(label, v["text"]), tags=(v["schwere"],))

        # Status-Label
        if not self.planning_result:
            self.check_status_label.config(text="Noch kein Plan erstellt.", foreground="black")
        elif fehler_n == 0 and hinweis_n == 0:
            self.check_status_label.config(text="✓ Alle Regeln eingehalten.", foreground="#2e7d32")
        elif fehler_n == 0:
            self.check_status_label.config(
                text=f"✓ Keine Fehler — {hinweis_n} Hinweis(e).", foreground="#b26a00")
        else:
            self.check_status_label.config(
                text=f"✗ {fehler_n} Fehler, {hinweis_n} Hinweis(e) — bitte prüfen!", foreground="#c62828")

        # Zeilen im Planungsergebnis einfärben (Fehler vor Hinweis)
        sev_by_row: Dict[int, str] = {}
        for v in violations:
            r = v["row"]
            if r is None:
                continue
            if v["schwere"] == "fehler":
                sev_by_row[r] = "fehler"
            elif sev_by_row.get(r) != "fehler":
                sev_by_row[r] = "hinweis"
        for idx, item in enumerate(self.result_tree.get_children()):
            sev = sev_by_row.get(idx)
            self.result_tree.item(item, tags=((f"check_{sev}",) if sev else ()))

        # Wochenlast-Panel
        self._update_week_load_panel()

        # Fairness-Metrik (auch nach manuellen Änderungen aktuell halten)
        if self._fairness_base_counts is not None and self.planning_result:
            self._fairness_after = self._combine_counts(
                self._fairness_base_counts, self._count_plan_assignments(self.planning_result)
            )
        self._update_fairness_metric_display()

    def _update_week_load_panel(self) -> None:
        """Aktualisiert die Wochenlast-Tabelle (Einsätze pro MA im aktuellen Plan)."""
        if not hasattr(self, 'week_load_tree'):
            return
        for item in self.week_load_tree.get_children():
            self.week_load_tree.delete(item)

        entries = self.planning_result
        if not entries:
            return

        loads: Dict[str, List[int]] = {}
        for i, e in enumerate(entries):
            w = 0 if i < 5 else 1
            vm = e.get("Vormittag", "").strip()
            for shift_type in ("Vormittag", "Nachmittag", "Support"):
                ma = e.get(shift_type, "").strip()
                if not ma or (shift_type == "Support" and ma == vm):
                    continue
                loads.setdefault(ma, [0, 0])
                loads[ma][w] += 1

        # Auch MA mit 0 Einsätzen anzeigen (damit Lücken sichtbar werden)
        for ma in self._get_all_employees():
            loads.setdefault(ma, [0, 0])

        totals = {ma: w[0] + w[1] for ma, w in loads.items()}
        if not totals:
            return
        max_total = max(totals.values())
        min_total = min(totals.values())

        for ma in sorted(loads.keys(), key=lambda m: (-totals[m], m)):
            tags = ()
            if max_total != min_total:
                if totals[ma] == max_total:
                    tags = ("max_load",)
                elif totals[ma] == min_total:
                    tags = ("min_load",)
            self.week_load_tree.insert(
                "", tk.END,
                values=(ma, loads[ma][0], loads[ma][1], totals[ma]),
                tags=tags
            )

    # -------------------- Anzeige & Export --------------------

    def display_results(self) -> None:
        """Zeigt die Planungsergebnisse in der Treeview an (optimiert)"""
        # Alte Einträge löschen
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)

        # Neue Einträge hinzufügen (angepasst an neue kompakte Spalten)
        for row in self.planning_result:
            # Wochentag abkürzen für kompaktere Darstellung
            tag_kurz = row["Wochentag"][:2] if row["Wochentag"] else ""
            self.result_tree.insert(
                "", tk.END,
                values=(row["Datum"], tag_kurz, row["Vormittag"],
                       row["Nachmittag"], row["Support"])
            )

        # Plan-Prüfung, Wochenlast und Fairness-Metrik aktualisieren
        self._run_plan_checks()

    def _on_result_double_click(self, event) -> None:
        """Handler für Doppelklick zum Bearbeiten einer Zelle"""
        # Aktives Eingabefeld schließen
        if self._edit_entry:
            self._cancel_edit()

        # Klick-Position ermitteln
        region = self.result_tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = self.result_tree.identify_column(event.x)
        item = self.result_tree.identify_row(event.y)

        if not item or not column:
            return

        # Spaltenindex (1-basiert): #1=Datum, #2=Tag, #3=VM, #4=NM, #5=Support
        col_idx = int(column.replace("#", ""))

        # Nur VM, NM, Support editierbar (Spalten 3, 4, 5)
        if col_idx < 3:
            return

        # Aktuelle Werte holen
        values = list(self.result_tree.item(item, "values"))
        current_value = values[col_idx - 1] if col_idx - 1 < len(values) else ""

        # Zellen-Bounding-Box ermitteln
        bbox = self.result_tree.bbox(item, column)
        if not bbox:
            return

        x, y, width, height = bbox

        # Entry-Widget erstellen
        self._edit_entry = tk.Entry(self.result_tree, width=width // 8)
        self._edit_entry.place(x=x, y=y, width=width, height=height)
        self._edit_entry.insert(0, current_value)
        self._edit_entry.select_range(0, tk.END)
        self._edit_entry.focus_set()

        # Metadaten speichern
        self._edit_item = item
        self._edit_col_idx = col_idx
        self._edit_row_idx = self.result_tree.index(item)

        # Event-Bindings
        self._edit_entry.bind("<Return>", lambda e: self._confirm_edit())
        self._edit_entry.bind("<Escape>", lambda e: self._cancel_edit())
        self._edit_entry.bind("<FocusOut>", lambda e: self._confirm_edit())

    def _confirm_edit(self) -> None:
        """Bestätigt die Bearbeitung und aktualisiert die Daten"""
        if not self._edit_entry:
            return

        new_value = self._edit_entry.get().strip()

        # Treeview aktualisieren
        values = list(self.result_tree.item(self._edit_item, "values"))
        values[self._edit_col_idx - 1] = new_value
        self.result_tree.item(self._edit_item, values=values)

        # planning_result aktualisieren
        col_map = {3: "Vormittag", 4: "Nachmittag", 5: "Support"}
        if self._edit_col_idx in col_map:
            self.planning_result[self._edit_row_idx][col_map[self._edit_col_idx]] = new_value
            self._plan_manually_edited = True  # Flag setzen bei nachträglicher Änderung

        self._cleanup_edit()
        # Nach jeder manuellen Änderung sofort neu prüfen (Regelverstöße sichtbar machen)
        self._run_plan_checks()

    def _cancel_edit(self) -> None:
        """Bricht die Bearbeitung ab"""
        self._cleanup_edit()

    def _cleanup_edit(self) -> None:
        """Räumt das Edit-Widget auf"""
        if self._edit_entry:
            self._edit_entry.destroy()
            self._edit_entry = None
        self._edit_item = None
        self._edit_col_idx = None
        self._edit_row_idx = None

    def _prepare_export_data(self) -> pd.DataFrame:
        """Bereitet Daten für Export vor"""
        rows = []
        for r in self.planning_result:
            try:
                datum_dt = datetime.strptime(r["Datum"], "%d.%m.%Y")
            except Exception:
                datum_dt = r["Datum"]
            rows.append({
                "Datum": datum_dt,
                "Wochentag": r["Wochentag"],
                "Vormittag": r["Vormittag"],
                "Nachmittag": r["Nachmittag"],
                "Support": r["Support"]
            })
        return pd.DataFrame(rows, columns=["Datum", "Wochentag", "Vormittag", "Nachmittag", "Support"])

    def _export_csv(self, filename: str, df: pd.DataFrame) -> None:
        """Exportiert als CSV"""
        df_copy = df.copy()
        if not df_copy.empty and isinstance(df_copy.loc[0, "Datum"], datetime):
            df_copy["Datum"] = df_copy["Datum"].dt.strftime("%d.%m.%Y")
        df_copy.to_csv(filename, index=False, encoding="utf-8-sig")
        messagebox.showinfo("Erfolg", f"CSV-Datei erfolgreich gespeichert: {filename}")

    def export_excel(self) -> None:
        """Exportiert die Planung nach Excel im Layout:
        Spalten: Datum | Wochentag | Vormittag (grün) | Nachmittag (grün) | Support (gelb)
        Samstag wird aufgeführt (ohne Planung), nach jedem Samstag eine graue Trennzeile.
        """
        if not self.planning_result:
            messagebox.showwarning("Warnung", "Keine Planung vorhanden! Bitte erst Planung erstellen.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not filename:
            return

        df = self._prepare_export_data()

        # CSV-Fallback ohne Formatierung
        if filename.lower().endswith(".csv"):
            self._export_csv(filename, df)
            return

        # Excel mit Formatierung
        try:
            self._export_excel_formatted(filename, df)
        except ImportError:
            # Fallback zu CSV wenn openpyxl nicht verfügbar
            csv_filename = filename.rsplit(".", 1)[0] + ".csv"
            self._export_csv(csv_filename, df)
            messagebox.showinfo(
                "Info",
                "openpyxl ist nicht installiert. CSV gespeichert: "
                f"{csv_filename}\nFür formatiertes Excel bitte installieren: pip install openpyxl"
            )
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Export: {e}")

    def _export_excel_formatted(self, filename: str, df: pd.DataFrame) -> None:
        """Exportiert formatiertes Excel mit Farben und Trennzeilen"""
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        # Farbpalette
        GREEN = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        YELLOW = PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid")
        GREY = PatternFill(start_color=COLOR_GREY, end_color=COLOR_GREY, fill_type="solid")

        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        with pd.ExcelWriter(filename, engine="openpyxl", date_format="DD.MM.YYYY") as writer:
            df.to_excel(writer, sheet_name="Notdienst_Planung", index=False)
            ws = writer.sheets["Notdienst_Planung"]

            # Spaltenbreiten setzen
            column_widths = [12, 14, 14, 14, 14]  # A..E
            for idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            # Header formatieren
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin

            # Datenzellen formatieren
            max_row = ws.max_row
            max_col = ws.max_column

            # Farbzuordnung für Spalten
            column_fills = {3: GREEN, 4: GREEN, 5: YELLOW}

            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = thin
                    cell.alignment = Alignment(horizontal=("left" if c in (1, 2) else "center"))

                    # Füllfarbe anwenden
                    if c in column_fills:
                        cell.fill = column_fills[c]

            # Trennzeilen nach Samstagen einfügen
            self._insert_saturday_separators(ws, df, GREY, thin, max_col)

        messagebox.showinfo("Erfolg", f"Excel-Datei erfolgreich gespeichert: {filename}")

    def _insert_saturday_separators(self, ws, df: pd.DataFrame, grey_fill, border, max_col: int) -> None:
        """Fügt graue Trennzeilen nach jedem Freitag ein (Wochentrenner)"""
        friday_rows = []
        for i, weekday in enumerate(df["Wochentag"].tolist(), start=2):  # +1 Header, +1 1-basiert
            if weekday == "Freitag":
                friday_rows.append(i)

        # Von unten nach oben einfügen um Zeilenindizes nicht zu verschieben
        # Nur nach dem ersten Freitag trennen (nicht nach dem letzten)
        for row_idx in sorted(friday_rows[:-1] if len(friday_rows) > 1 else [], reverse=True):
            insert_row = row_idx + 1
            ws.insert_rows(insert_row)
            for c in range(1, max_col + 1):
                cell = ws.cell(row=insert_row, column=c)
                cell.fill = grey_fill
                cell.border = border


def main():
    root = tk.Tk()
    app = ShiftPlanner(root)
    root.mainloop()


if __name__ == "__main__":
    main()
